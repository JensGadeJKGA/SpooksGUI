
# -*- coding: utf-8 -*-
#"""
#Created on Sat Oct 24 11:45:53 2020

#@author: EMBT, JKGA
#"""




####### This script intends to calculate spookswat analyses from an excel input file
####### structured as SPOOKSINPUT.xlsx. 
## Vers 2.1
# Prepared: EMBT, 10.10.2023
# Checked: SEMS, CHHJ
## Vers 1.0
# Prepared: EMBT, 06.07.2020
# Checked: EMSS, 06.07.2020


### Imports
from Utils import Utils

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf
import subprocess
import io
import os
import tkinter as tk
import tkinter.scrolledtext as tkst
import contextlib
import PyPDF2
import os
import subprocess
import time

from fpdf import FPDF
from getpass import getuser
from datetime import datetime
from openpyxl import load_workbook
from tkinter import ttk
from tkinter import filedialog
from scipy.interpolate import interp1d
from shutil import copyfile


################ Version ##################

Version = '2.1'

################ Input file ID ############

InputFileIDGUI = 'A3'


############### IMPORT SHEET PILE WALL PLUG IN ###################

from Steel_Sheet_Pile_Wall import steel_sheet_pile_implementer


def log_usage():

    import getpass
    import socket
    try:
        log_path = r"O:\Organisation\DK_1551\Diverse\SpooksLog"
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        user = getpass.getuser()
        pc_name = socket.gethostname()
        log = os.path.join(log_path, "log.log")
        with open(log, "a") as f:
            f.write(f"{timestamp}\t{user.upper()}\t{pc_name.upper()}\n")
    except:
        pass

################### FUNCTIONS ####################################

def InputFileIDChecker(InputFileID):
    
    InputFileStatus = None
    
    if InputFileID == InputFileIDGUI:
        
        InputFileStatus = 'OK'
        
    else:
        
        stat.configure(text = 'Input file version does not match GUI version')
        
    return InputFileStatus


        
def TemporaryWorkingDirectory():
    

    user = "MLHU"
    
    ## output path
    TemporaryPath = os.path.join(r'C:\Users', user)
    TemporaryPath = os.path.join(TemporaryPath, r'AppData\Local\Temp\2')
    
    if not os.path.exists(TemporaryPath): ## if directory does not exist -> create it
        os.makedirs(TemporaryPath)
        
    return TemporaryPath



def ImportExcel(input_path):
    
    ## Libraries

    
    
    #### Loading excel workbook
    xlsx_filename=input_path
    with open(xlsx_filename, "rb") as f:
        in_mem_file = io.BytesIO(f.read())
    
    wb = load_workbook(in_mem_file, data_only=True)
            
    ## Sheets
    INFO =  wb['INFO']
    GeneralInfo = wb['General_information']
    Stratification = wb['Stratification']
    Wall = wb['Wall']
    Water = wb['Water']
    Add_pres = wb['Additional_pressure_profiles']
    Analyses = wb['Analyses']
    LoadComb = wb['Load_combinations']
    SheetPileAddOn = wb['Addon - Sheet Pile Wall']
    
    
    ########### 3.1. LOADING DATA SHEETS TO PANDAS DATAFRAMES #########

    
    data_array = [INFO['A1':'H28'], GeneralInfo['A3':'B7'], Stratification['A4':'K289'], Wall['A3':'B5'], Water['A3':'A5'], Add_pres['A3':'C114'], Analyses['A3':'AA56'], LoadComb['A3':'M42'], SheetPileAddOn['A1':'G30']]
    docready_array = Utils.data_rows_arr(data_array)

    ###### This returns an array with all the data post processing. So docready_array[0] is INFO, docready_array[1] is stratification etc.
    INFO = pd.DataFrame(docready_array[0])
    Stratification = docready_array[1]
    InputFileID = str(INFO.iloc[27,1])

    ImportData = {'InputFileID': InputFileID, 
                  'GeneralInfo': docready_array[1],
                  'Stratification': docready_array[2],
                  'Wall': docready_array[3],
                  'Water': docready_array[4],
                  'AddPress': docready_array[5],
                  'Analyses': docready_array[6],
                  'LoadComb': docready_array[7],
                  'SheetPileAddOn': docready_array[8]}
    

    return ImportData









def GenerateSoilProfiles(Stratification):
    
    SoilProfiles = {'SP1': {'Back': {'Slope': None,'Layers': []}, 'Front': {'Slope': None,'Layers': []}},
                    'SP2': {'Back': {'Slope': None,'Layers': []}, 'Front': {'Slope': None,'Layers': []}},
                    'SP3': {'Back': {'Slope': None,'Layers': []}, 'Front': {'Slope': None,'Layers': []}},
                    'SP4': {'Back': {'Slope': None,'Layers': []}, 'Front': {'Slope': None,'Layers': []}},
                    'SP5': {'Back': {'Slope': None,'Layers': []}, 'Front': {'Slope': None,'Layers': []}},
                    'SP6': {'Back': {'Slope': None,'Layers': []}, 'Front': {'Slope': None,'Layers': []}},
                    'SP7': {'Back': {'Slope': None,'Layers': []}, 'Front': {'Slope': None,'Layers': []}},
                    'SP8': {'Back': {'Slope': None,'Layers': []}, 'Front': {'Slope': None,'Layers': []}},
                    'SP9': {'Back': {'Slope': None,'Layers': []}, 'Front': {'Slope': None,'Layers': []}},
                    'SP10': {'Back': {'Slope': None,'Layers': []}, 'Front': {'Slope': None,'Layers': []}}}


    
    ################### FRONT SOILS
    for i in range(10):
        sp = "SP"+str(i+1)
        Utils.AppendToSoilProfiles(Utils.soilprofiles(sp, Stratification, SoilProfiles, [29*i,5], i))

    #### Soil profile 1
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP1', Stratification, SoilProfiles, [0,5], 0))

    #### Soil profile 2
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP2', Stratification, SoilProfiles, [28,5], 1))
    
    #### Soil profile 3
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP3', Stratification, SoilProfiles, [56,5], 2))
    
    #### Soil profile 4
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP4', Stratification, SoilProfiles, [85,5], 3))
    
    #### Soil profile 5
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP5', Stratification, SoilProfiles, [114,5], 4))
    
    #### Soil profile 6
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP6', Stratification, SoilProfiles, [143,5], 5))        
                
    #### Soil profile 7
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP7', Stratification, SoilProfiles, [172,5], 6))
    
    #### Soil profile 8
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP8', Stratification, SoilProfiles, [201,5], 7))

    #### Soil profile 9
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP9', Stratification, SoilProfiles, [230,5], 8))
    
    #### Soil profile 10
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP8', Stratification, SoilProfiles, [259,5], 8))


    # ############# BACK SOILS
    
    for i in range(10):
        sp = "SP"+str(i+1)
        Utils.AppendToSoilProfiles(Utils.soilprofiles(sp, Stratification, SoilProfiles, [29*i,2], i))

    #### Soil profile 1
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP1', Stratification, SoilProfiles, [0,2], 0))
    
    #### Soil profile 2
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP2', Stratification, SoilProfiles, [28,2], 1))
    
    #### Soil profile 3
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP3', Stratification, SoilProfiles, [56,2], 2))
    
    #### Soil profile 4
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP4', Stratification, SoilProfiles, [85,2], 3))
    
    #### Soil profile 5
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP5', Stratification, SoilProfiles, [114,2], 4))

    #### Soil profile 6
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP6', Stratification, SoilProfiles, [143,2], 6))
    
    #### Soil profile 7
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP7', Stratification, SoilProfiles, [172,2], 7))

    #### Soil profile 8
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP8', Stratification, SoilProfiles, [201,2], 8))

    #### Soil profile 9
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP9', Stratification, SoilProfiles, [231,2], 9))
    
    #### Soil profile 10
    #Utils.AppendToSoilProfiles(Utils.soilprofiles('SP9', Stratification, SoilProfiles, [259,2], 10))
    
    return SoilProfiles
    
    
    
def WaterDensity(Water):
    ######## WATER DENSITY
    
    gamma_water = float(format(float(Water.iloc[1,0]), '.2f'))
    
    return gamma_water



def WallParameters(Wall):

    
    ############# WALL PARAMETERS
    ### zT (Top wall)
    
    zT = float(format(float(Wall.iloc[1,0]), '.2f'))
    
    ### Mass of wall (for vertical equilibrium)
    wall_mass = float(Wall.iloc[1,1])
    
    WallParams = {'zT': zT,
                  'Mass': wall_mass}
    
    return WallParams




def GenerateAddPressProfiles(Add_pres):
    
    AdditionalPressures = {'AP1':{'z': [],'ez':[]},
                           'AP2': {'z': [],'ez':[]},
                           'AP3': {'z': [],'ez':[]},
                           'AP4': {'z': [],'ez':[]},
                           'AP5': {'z': [],'ez':[]},
                           'AP6': {'z': [],'ez':[]},
                           'AP7': {'z': [],'ez':[]},
                           'AP8': {'z': [],'ez':[]},
                           'AP9': {'z': [],'ez':[]},
                           'AP10': {'z': [],'ez':[]}}
    
    
    ############# ADDITTIONAL PRESSURE PROFILES
    
    #### AP1
    for i in range(9):
        AdditionalPressures = Utils.AddPressProfiles(i+1, Add_pres, AdditionalPressures)
    
    return AdditionalPressures




def GeneratePartialCoefficientDictionary(LoadComb):
    
    
    LoadCombinations = {'CC2': {},
                        'CC3': {}}
              


    ### Find CC2 partial safety factors
    index_CC2 = np.where((LoadComb.iloc[:,0]) == 'CC2')
    index_CC2 = index_CC2[0][0]
    index_CC3 = np.where((LoadComb.iloc[:,0]) == 'CC3')
    index_CC3 = index_CC3[0][0]
    
    for i in range(index_CC2+2,index_CC3):
        
        if LoadComb.iloc[i,0] != None:
            
            Loadcombination = LoadComb.iloc[i,0]

                        
            PartialSafetyFactors = {'f_gamf': float(LoadComb.iloc[i,1]),
                                    'f_qf':   float(LoadComb.iloc[i,2]),
                                    'f_cf':   float(LoadComb.iloc[i,3]),
                                    'f_cuf':  float(LoadComb.iloc[i,4]),
                                    'f_phif': float(LoadComb.iloc[i,5]),
                                    'f_wat':  float(LoadComb.iloc[i,6]),
                                    'f_AP':   float(LoadComb.iloc[i,7]),
                                    'f_gamb': float(LoadComb.iloc[i,8]),
                                    'f_qb':   float(LoadComb.iloc[i,9]),
                                    'f_cb':   float(LoadComb.iloc[i,10]),
                                    'f_cub':  float(LoadComb.iloc[i,11]),
                                    'f_phib': float(LoadComb.iloc[i,12])}
            
            
            LoadCombinations.get('CC2')[str(Loadcombination)] = PartialSafetyFactors
            
        
    ### Find CC3 partial safety factors
    for i in range(index_CC3+2,len(LoadComb.iloc[:,0])):
        
        if LoadComb.iloc[i,0] != None:
            
            Loadcombination = LoadComb.iloc[i,0]

                        
            PartialSafetyFactors = {'f_gamf': float(LoadComb.iloc[i,1]),
                                    'f_qf':   float(LoadComb.iloc[i,2]),
                                    'f_cf':   float(LoadComb.iloc[i,3]),
                                    'f_cuf':  float(LoadComb.iloc[i,4]),
                                    'f_phif': float(LoadComb.iloc[i,5]),
                                    'f_wat':  float(LoadComb.iloc[i,6]),
                                    'f_AP':   float(LoadComb.iloc[i,7]),
                                    'f_gamb': float(LoadComb.iloc[i,8]),
                                    'f_qb':   float(LoadComb.iloc[i,9]),
                                    'f_cb':   float(LoadComb.iloc[i,10]),
                                    'f_cub':  float(LoadComb.iloc[i,11]),
                                    'f_phib': float(LoadComb.iloc[i,12])}

            
            LoadCombinations.get('CC3')[str(Loadcombination)] = PartialSafetyFactors
            
                                
                        
    return LoadCombinations


def GenerateSheetPileAddOnInput(SheetPileAddOn):
    
    UseAddOn = SheetPileAddOn.iloc[0,4]
    LimitState = SheetPileAddOn.iloc[3,6]
    ControlClass = SheetPileAddOn.iloc[4,6]
    KFI = SheetPileAddOn.iloc[5,6]
    Optimize = SheetPileAddOn.iloc[8,6]
    MaxUtilization = SheetPileAddOn.iloc[9,6]
    fyk = SheetPileAddOn.iloc[10,6]
    BetaB = SheetPileAddOn.iloc[11,6]
    BetaD = SheetPileAddOn.iloc[12,6]
    DesignLife = SheetPileAddOn.iloc[15,6]
    tCor = list(SheetPileAddOn.iloc[17:27,6])
    tCorLevel = list(SheetPileAddOn.iloc[17:27,0])
    SoilDeposit = SheetPileAddOn.iloc[29,6]
    
    
    
    SheetPileAddOnInput = {'UseAddOn': UseAddOn,
                           'LimitState': LimitState,
                           'ControlClass': ControlClass,
                           'KFI': KFI,
                           'Optimize': Optimize,
                           'MaxUtilization': MaxUtilization,
                           'fyk': fyk,
                           'BetaB': BetaB,
                           'BetaD': BetaD,
                           'DesignLife': DesignLife,
                           'tCor': tCor,
                           'tCorLevel': tCorLevel,
                           'SoilDeposit': SoilDeposit}
    
    return SheetPileAddOnInput
    
    
    
    

def AnalysesRange(Analyses):
    
    ### The script checks if data is entered in all relevant analysis columns
    null0 = np.amin(np.where(pd.isnull(Analyses.iloc[:,0]))) ### Script aims for the first cell in which the user hasn't entered Subject text
    null1 = np.amin(np.where(pd.isnull(Analyses.iloc[:,1]))) ### Script aims for the first cell of which the user hasn't entered Soil profile (SP)
    null2 = np.amin(np.where(pd.isnull(Analyses.iloc[:,2]))) ### Script aims for the first cell in which the user hasn't entered "Drained" or "Undrained"
    null3 = np.amin(np.where(pd.isnull(Analyses.iloc[:,3]))) ### Script aims for the first cell of which the user hasn't entered water front level (w_f)
    null4 = np.amin(np.where(pd.isnull(Analyses.iloc[:,4]))) ### Script aims for the first cell of which the user hasn't entered water back level (w_b)
    null5 = np.amin(np.where(pd.isnull(Analyses.iloc[:,5]))) ### Script aims for the first cell of which the user hasn't entered load combination
    null6 = np.amin(np.where(pd.isnull(Analyses.iloc[:,6]))) ### Script aims for the first cell of which the user hasn't entered consequence class
    null7 = np.amin(np.where(pd.isnull(Analyses.iloc[:,7]))) ### Script aims for the first cell of which the user hasn't entered alpha value
    null8 = np.amin(np.where(pd.isnull(Analyses.iloc[:,8]))) ### Script aims for the first cell of which the user hasn't entered front load value (0,00 okay)
    null9 = np.amin(np.where(pd.isnull(Analyses.iloc[:,9]))) ### Script aims for the first cell of which the user hasn't entered back load value (0,00 okay)
    null10 = np.amin(np.where(pd.isnull(Analyses.iloc[:,10]))) ### Script aims for the first cell of which the user hasn't entered zR value
    ### The user does not need to enter Additional pressure value (no. 11 missing)
    null12 = np.amin(np.where(pd.isnull(Analyses.iloc[:,12]))) ### Script aims for the first cell of which the user hasn't entered iA value
    null13 = np.amin(np.where(pd.isnull(Analyses.iloc[:,13]))) ### Script aims for the first cell of which the user hasn't entered iA value
    null14 = np.amin(np.where(pd.isnull(Analyses.iloc[:,14]))) ### Script aims for the first cell of which the user hasn't entered iC value
    
    ### Overall first row (analysis row) with incomplete data
    index_maxAnalysis = min(null0, null1, null2, null3, null4, null5,
                null6, null7, null8, null9, null10, null12, null13, null14)
    
    
    index_minAnalysis = 2 # Row of first analysis
    
    
    RangeOfAnalyses = {'MinAnalysis': index_minAnalysis,
                       'MaxAnalysis': index_maxAnalysis}
    
    
    
    return RangeOfAnalyses

    
def AddSoilToAnalysis(GeneratedAnalyses,SoilProfiles):
    
    ## Loop through all generated analyses and append stratigraphy input
    for Analysis in GeneratedAnalyses:
        
        SoilProfile = Analysis.get('SoilProfile')
        
        ## Find analysis soil profile in SoilProfiles
        SP = SoilProfiles.get(SoilProfile)
        
        ## Append slope back
        Analysis['SlopeBack'] = SP.get('Back').get('Slope')
        
        ## Append slope front
        Analysis['SlopeFront'] = SP.get('Front').get('Slope')
        
        ## Append soil profile properties (back)
        for SoilLayer in SP.get('Back').get('Layers'):
        
                Analysis.get('SoilLayersBack').append(SoilLayer)
                
        ## Append soil profile properties (front)
        for SoilLayer in SP.get('Front').get('Layers'):
        
                Analysis.get('SoilLayersFront').append(SoilLayer)
        


def AddPressureToAnalysis(GeneratedAnalyses,AdditionalPressures):
    
    ## Loop through all generated analyses and append stratigraphy input
    for Analysis in GeneratedAnalyses:
        
        APProfile = Analysis.get('AddPressureProfile')
        
        ## If any of the possible additional pressure profiles is specified
        if APProfile != None:
            
            ## Additional pressure profile
            APProfile = AdditionalPressures.get(APProfile)
        
            ## Append AP levels
            for Level in APProfile.get('z'):
            
                    Analysis.get('AddPress_z').append(Level)
            
            ## Append AP pressures
            for Pressure in APProfile.get('ez'):
            
                    Analysis.get('AddPress_ez').append(Pressure)
                    


def AddDesignParameters(GeneratedAnalyses,LoadComb):
    
    LoadCombinations = GeneratePartialCoefficientDictionary(LoadComb)
    
    
    for Analysis in GeneratedAnalyses:
    
        ConsequenceClass = Analysis.get('ConsequenceClass')
        LoadCombination = Analysis.get('LoadCombination')
        Alpha = float(Analysis.get('Alpha'))
        SoilLayersBack = Analysis.get('SoilLayersBack')
        SoilLayersFront = Analysis.get('SoilLayersFront')
        LoadFront = Analysis.get('LoadFront')
        LoadBack = Analysis.get('LoadBack')
        WaterDensity = Analysis.get('WaterDensity')
        AddPress_ez = Analysis.get('AddPress_ez')
        
        
        DesignSoilLayersBack = []
        DesignSoilLayersFront = []
        
        PartialSafetyFactors = LoadCombinations.get(ConsequenceClass).get(LoadCombination)
        
        
        
        ###################### SOILS #######################################
        
        
        ## Generate design soil layers (back)
        for SoilLayer in SoilLayersBack:
            
            TopLayer = float(SoilLayer.get('TopLayer'))
            Gamma_d = float(SoilLayer.get('Gamma_d'))
            Gamma_m = float(SoilLayer.get('Gamma_m'))
            cu = float(SoilLayer.get('cu'))
            c = float(SoilLayer.get('c'))
            phi = float(SoilLayer.get('phi'))
            i = float(SoilLayer.get('i'))
            r = float(SoilLayer.get('r'))
            Description = SoilLayer.get('Description')
            KeepDrained = SoilLayer.get('KeepDrained')
            
            
            DesignSoilLayer = {'TopLayer': TopLayer,
                               'Gamma_d':  Gamma_d / (PartialSafetyFactors.get('f_gamb')**Alpha),
                               'Gamma_m':  Gamma_m / (PartialSafetyFactors.get('f_gamb')**Alpha),
                               'cu':       cu / (PartialSafetyFactors.get('f_cub')**Alpha),
                               'c':        c / (PartialSafetyFactors.get('f_cb')**Alpha),
                               'phi':      np.degrees(np.arctan(np.tan(np.radians(phi))/(PartialSafetyFactors.get('f_phib')**Alpha))),
                               'i':        i,
                               'r':        r,
                               'Description': Description,
                               'KeepDrained': KeepDrained}
            
            DesignSoilLayersBack.append(DesignSoilLayer)
            
        Analysis['DesignSoilLayersBack'] = DesignSoilLayersBack
        
        
        ## Generate design soil layers (front)
        for SoilLayer in SoilLayersFront:
            
            TopLayer = float(SoilLayer.get('TopLayer'))
            Gamma_d = float(SoilLayer.get('Gamma_d'))
            Gamma_m = float(SoilLayer.get('Gamma_m'))
            cu = float(SoilLayer.get('cu'))
            c = float(SoilLayer.get('c'))
            phi = float(SoilLayer.get('phi'))
            i = float(SoilLayer.get('i'))
            r = float(SoilLayer.get('r'))
            Description = SoilLayer.get('Description')
            KeepDrained = SoilLayer.get('KeepDrained')
            
            
            DesignSoilLayer = {'TopLayer': TopLayer,
                               'Gamma_d':  Gamma_d / (PartialSafetyFactors.get('f_gamf')**Alpha),
                               'Gamma_m':  Gamma_m / (PartialSafetyFactors.get('f_gamf')**Alpha),
                               'cu':       cu / (PartialSafetyFactors.get('f_cuf')**Alpha),
                               'c':        c / (PartialSafetyFactors.get('f_cf')**Alpha),
                               'phi':      np.degrees(np.arctan(np.tan(np.radians(phi))/(PartialSafetyFactors.get('f_phif')**Alpha))),
                               'i':        i,
                               'r':        r,
                               'Description': Description,
                               'KeepDrained': KeepDrained}
            
            DesignSoilLayersFront.append(DesignSoilLayer)
        
        Analysis['DesignSoilLayersFront'] = DesignSoilLayersFront
        
        
        
        
        ###################### ADDITIONAL PRESSURE ##########################  
        
        DesignAddPress_ez = []
        
        for ez in AddPress_ez:
            
            DesignAddPress_ez.append(float(ez)*float(PartialSafetyFactors.get('f_AP')))
            
        
        Analysis['DesignAddPress_ez'] = DesignAddPress_ez
        
        
        ################## LOADS AND WATER DENSITY ###########################
        
        
        DesignLoadFront = float(LoadFront)*float(PartialSafetyFactors.get('f_qf'))
        DesignLoadBack = float(LoadBack)*float(PartialSafetyFactors.get('f_qb'))
        DesignWaterDensity = float(WaterDensity)*float(PartialSafetyFactors.get('f_wat'))
        
        Analysis['DesignLoadFront'] = DesignLoadFront
        Analysis['DesignLoadBack'] = DesignLoadBack
        Analysis['DesignWaterDensity'] = DesignWaterDensity
        Analysis['PartialSafetyFactors'] = PartialSafetyFactors
    
    

def GenerateAnalyses(input_path):
    
    ## Importing Excel file
    ImportData = ImportExcel(input_path)
    
    InputFileID = ImportData.get('InputFileID')
    Analyses = ImportData.get('Analyses')
    Water = ImportData.get('Water')
    Wall = ImportData.get('Wall')
    Stratification = ImportData.get('Stratification')
    SoilProfiles = GenerateSoilProfiles(Stratification)
    Add_pres = ImportData.get('AddPress')
    AdditionalPressures = GenerateAddPressProfiles(Add_pres)
    LoadComb = ImportData.get('LoadComb')
    SheetPileAddOn = ImportData.get('SheetPileAddOn')
    
    
    ## Check if input file version matches GUI version
    InputFileStatus = InputFileIDChecker(InputFileID)
    
    if InputFileStatus == 'OK':
        print("OK")
    
        ## List holding analyses (output)
        GeneratedAnalyses = []
        ## Initial analysis number
        AnalysisNo = 0
        ## Parent analysis (Referring to Excel input file)
        ParentAnalysis = 0
        
        ## Project
        Project = str(ImportData.get('GeneralInfo').iloc[0,1])
        ## Initials
        Initials = str(ImportData.get('GeneralInfo').iloc[2,1])
        ## Checker
        Checker = str(ImportData.get('GeneralInfo').iloc[3,1])
        ## Approver
        Approver = str(ImportData.get('GeneralInfo').iloc[4,1])
        ## Water density
        gamma_water = WaterDensity(Water)
        ## Wall parameters
        WallParams = WallParameters(Wall)
        WallMass = WallParams.get('Mass')
        TopWall = WallParams.get('zT')
        
        ### Sheet pile add on
        SheetPileAddOnInput = GenerateSheetPileAddOnInput(SheetPileAddOn)
    
        
        
        ### Ranges of analyses
        RangeOfAnalyses = AnalysesRange(Analyses)
        print(RangeOfAnalyses)
        MinAnalysis = RangeOfAnalyses.get('MinAnalysis')
        MaxAnalysis = RangeOfAnalyses.get('MaxAnalysis')
    
    
        ### Defining correct number of decimals for integers and floating numbers
        for Analysis in range(MinAnalysis,MaxAnalysis):
            for i in range(0, len(Analyses.iloc[Analysis,:])):
            
                if isinstance(Analyses.iloc[Analysis,i], int) == True or isinstance(Analyses.iloc[Analysis,i], float) == True:  ## add relevant number of decimals if 'i' is integer
                    Analyses.iloc[Analysis,i] = float(format(Analyses.iloc[Analysis,i], '.2f'))
                            
        
        ### Generating analyses
        for Analysis in range(MinAnalysis,MaxAnalysis):
            
            
            ### King Post Wall data
            zB = Analyses.iloc[Analysis,21]
            WD = Analyses.iloc[Analysis,22]
            CC = Analyses.iloc[Analysis,23]
            
            if isinstance(zB,(int,float)) == False or isinstance(WD,(int,float)) == False or isinstance(CC,(int,float)) == False:
                
                zB = None
                WD = None
                CC = None
            
            
            ## If anchor input is present
            if isinstance(Analyses.iloc[Analysis,16],(int,float)):
                print(1)
                AInclination = Analyses.iloc[Analysis,19]
                
                PrescrbAnchorForce = Analyses.iloc[Analysis,20]
                
                ## If no anchor inclination data is entered inclination is set to 0.00 degrees
                if isinstance(AInclination,(int,float)) == False:
                    
                    AInclination = 0.00
                    
                ## If no prescribed anchor force data is entered inclination is set to 0.00 degrees
                if isinstance(PrescrbAnchorForce,(int,float)) == False:
                    
                    PrescrbAnchorForce = 0.00
                
                ## If input is setup for anchor level range analysis
                if Analyses.iloc[Analysis,16] != Analyses.iloc[Analysis,17] and Analyses.iloc[Analysis,18] != None: ## checking anchor data (Alvl_min, Alvl_max and Astep)
                    print(1.1)
                    
                    ## Generate analysis for all anchor level steps
                    for AnchorLevel in np.arange(Analyses.iloc[Analysis,16],
                                                 Analyses.iloc[Analysis,17]+Analyses.iloc[Analysis,18],
                                                 Analyses.iloc[Analysis,18]):
            
            
                        Analysis_dict = {'AnalysisNo': AnalysisNo,
                                         'ParentAnalysis': ParentAnalysis,
                                         'Project': Project,
                                         'Initials': Initials,
                                         'Checker': Checker,
                                         'Approver': Approver,
                                         'Subject': Analyses.iloc[Analysis,0],
                                         'SoilProfile': Analyses.iloc[Analysis,1],
                                         'SlopeBack': None,
                                         'SlopeFront': None,
                                         'SoilLayersFront': [],
                                         'SoilLayersBack': [],
                                         'zT': float(TopWall),
                                         'WallMass': float(WallMass),
                                         'WaterDensity': float(gamma_water),
                                         'AddPressureProfile': Analyses.iloc[Analysis,11],
                                         'AddPress_z': [],
                                         'AddPress_ez': [],
                                         'State': Analyses.iloc[Analysis,2],
                                         'WaterLevelFront': float(Analyses.iloc[Analysis,3]),
                                         'WaterLevelBack': float(Analyses.iloc[Analysis,4]),
                                         'WDiff': abs(float(Analyses.iloc[Analysis,4])-float(Analyses.iloc[Analysis,4])),
                                         'LoadCombination': Analyses.iloc[Analysis,5],
                                         'ConsequenceClass': Analyses.iloc[Analysis,6],
                                         'Alpha': float(Analyses.iloc[Analysis,7]),
                                         'LoadFront': float(Analyses.iloc[Analysis,8]),
                                         'LoadBack': float(Analyses.iloc[Analysis,9]),
                                         'LevelLoadBack': float(Analyses.iloc[Analysis,10]),
                                         'AxialWallLoad': float(Analyses.iloc[Analysis,12]),
                                         'iA': float(Analyses.iloc[Analysis,13]),
                                         'iB': float(Analyses.iloc[Analysis,14]),
                                         'iC': float(Analyses.iloc[Analysis,15]),
                                         'AnchorLevel': float(AnchorLevel),
                                         'AnchorInclination': float(AInclination),
                                         'PrescrbAnchorForce': float(PrescrbAnchorForce),
                                         'zB': zB,
                                         'WD': WD,
                                         'CC': CC,
                                         'KN': Analyses.iloc[Analysis,24],
                                         'KP': Analyses.iloc[Analysis,25],
                                         'SC': Analyses.iloc[Analysis,26],
                                         'SheetPileAddOnInput': SheetPileAddOnInput}
                        
                        print(Analysis_dict)
                        
                        ## Append to analyses list
                        GeneratedAnalyses.append(Analysis_dict)
                                
                        ## Update analysis number
                        AnalysisNo = AnalysisNo + 1
                          
    
                else:
                    print(1.2)
                    Analysis_dict = {'AnalysisNo': AnalysisNo,
                                     'ParentAnalysis': ParentAnalysis,
                                     'Project': Project,
                                     'Initials': Initials,
                                     'Checker': Checker,
                                     'Approver': Approver,
                                     'Subject': Analyses.iloc[Analysis,0],
                                     'SoilProfile': Analyses.iloc[Analysis,1],
                                     'SlopeBack': None,
                                     'SlopeFront': None,
                                     'SoilLayersFront': [],
                                     'SoilLayersBack': [],
                                     'zT': float(TopWall),
                                     'WallMass': float(WallMass),
                                     'WaterDensity': float(gamma_water),
                                     'AddPressureProfile': Analyses.iloc[Analysis,11],
                                     'AddPress_z': [],
                                     'AddPress_ez': [],
                                     'State': Analyses.iloc[Analysis,2],
                                     'WaterLevelFront': float(Analyses.iloc[Analysis,3]),
                                     'WaterLevelBack': float(Analyses.iloc[Analysis,4]),
                                     'WDiff': abs(float(Analyses.iloc[Analysis,4])-float(Analyses.iloc[Analysis,4])),
                                     'LoadCombination': Analyses.iloc[Analysis,5],
                                     'ConsequenceClass': Analyses.iloc[Analysis,6],
                                     'Alpha': float(Analyses.iloc[Analysis,7]),
                                     'LoadFront': float(Analyses.iloc[Analysis,8]),
                                     'LoadBack': float(Analyses.iloc[Analysis,9]),
                                     'LevelLoadBack': float(Analyses.iloc[Analysis,10]),
                                     'AxialWallLoad': float(Analyses.iloc[Analysis,12]),
                                     'iA': float(Analyses.iloc[Analysis,13]),
                                     'iB': float(Analyses.iloc[Analysis,14]),
                                     'iC': float(Analyses.iloc[Analysis,15]),
                                     'AnchorLevel': float(Analyses.iloc[Analysis,16]),
                                     'AnchorInclination': float(AInclination),
                                     'PrescrbAnchorForce': float(PrescrbAnchorForce),
                                     'zB': zB,
                                     'WD': WD,
                                     'CC': CC,
                                     'KN': Analyses.iloc[Analysis,24],
                                     'KP': Analyses.iloc[Analysis,25],
                                     'SC': Analyses.iloc[Analysis,26],
                                     'SheetPileAddOnInput': SheetPileAddOnInput}
                    print(Analysis_dict)
                    ## Append to analyses list
                    GeneratedAnalyses.append(Analysis_dict)
                            
                    ## Update analysis number
                    AnalysisNo = AnalysisNo + 1
            else:        
                print(2)
                Analysis_dict = {'AnalysisNo': AnalysisNo,
                                 'ParentAnalysis': ParentAnalysis,
                                 'Project': Project,
                                 'Initials': Initials,
                                 'Checker': Checker,
                                 'Approver': Approver,
                                 'Subject': Analyses.iloc[Analysis,0],
                                 'SoilProfile': Analyses.iloc[Analysis,1],
                                 'SlopeBack': None,
                                 'SlopeFront': None,
                                 'SoilLayersFront': [],
                                 'SoilLayersBack': [],
                                 'zT': float(TopWall),
                                 'WallMass': float(WallMass),
                                 'WaterDensity': float(gamma_water),
                                 'AddPressureProfile': Analyses.iloc[Analysis,11],
                                 'AddPress_z': [],
                                 'AddPress_ez': [],
                                 'State': Analyses.iloc[Analysis,2],
                                 'WaterLevelFront': float(Analyses.iloc[Analysis,3]),
                                 'WaterLevelBack': float(Analyses.iloc[Analysis,4]),
                                 'WDiff': abs(float(Analyses.iloc[Analysis,4])-float(Analyses.iloc[Analysis,4])),
                                 'LoadCombination': Analyses.iloc[Analysis,5],
                                 'ConsequenceClass': Analyses.iloc[Analysis,6],
                                 'Alpha': float(Analyses.iloc[Analysis,7]),
                                 'LoadFront': float(Analyses.iloc[Analysis,8]),
                                 'LoadBack': float(Analyses.iloc[Analysis,9]),
                                 'LevelLoadBack': float(Analyses.iloc[Analysis,10]),
                                 'AxialWallLoad': float(Analyses.iloc[Analysis,12]),
                                 'iA': float(Analyses.iloc[Analysis,13]),
                                 'iB': float(Analyses.iloc[Analysis,14]),
                                 'iC': float(Analyses.iloc[Analysis,15]),
                                 'AnchorLevel': None,
                                 'AnchorInclination': None,
                                 'PrescrbAnchorForce': None,
                                 'zB': zB,
                                 'WD': WD,
                                 'CC': CC,
                                 'KN': Analyses.iloc[Analysis,24],
                                 'KP': Analyses.iloc[Analysis,25],
                                 'SC': Analyses.iloc[Analysis,26],
                                 'SheetPileAddOnInput': SheetPileAddOnInput}
            
                print(Analysis_dict)
                        
                ## Append to analyses list
                GeneratedAnalyses.append(Analysis_dict)
                            
                ## Update analysis number
                AnalysisNo = AnalysisNo + 1
            
            ## Update parent analysis number
            ParentAnalysis = ParentAnalysis + 1    
        
        
        ## Append soil layers data to analysis dictionary
        AddSoilToAnalysis(GeneratedAnalyses,SoilProfiles)
        
        ## Append soil layers data to analysis dictionary
        AddPressureToAnalysis(GeneratedAnalyses,AdditionalPressures)
        
        ## Append design soil layers
        AddDesignParameters(GeneratedAnalyses,LoadComb)
        
        
    
        return GeneratedAnalyses
    


def GenerateSPOOKSInputFile(Analysis):
    
    Project = Analysis.get('Project')
    Initials = Analysis.get('Initials')
    Checker = Analysis.get('Checker')
    Approver = Analysis.get('Approver')
    Subject = Analysis.get('Subject')
    SlopeFront = Analysis.get('SlopeFront')
    SlopeBack = Analysis.get('SlopeBack')
    LoadFront = Analysis.get('DesignLoadFront')
    LoadBack = Analysis.get('DesignLoadBack')
    WaterLevelBack = Analysis.get('WaterLevelBack')
    WaterLevelFront = Analysis.get('WaterLevelFront')
    WaterDensity = Analysis.get('WaterDensity')
    State = Analysis.get('State')
    DesignSoilLayersFront = Analysis.get('DesignSoilLayersFront')
    DesignSoilLayersBack = Analysis.get('DesignSoilLayersBack')
    iA = Analysis.get('iA')
    iB = Analysis.get('iB')
    iC = Analysis.get('iC')
    zT = Analysis.get('zT')
    zR = Analysis.get('LevelLoadBack')
    AnchorLevel = Analysis.get('AnchorLevel')
    PrescrbAnchorForce = Analysis.get('PrescrbAnchorForce')
    zB = Analysis.get('zB')
    WD = Analysis.get('WD')
    CC = Analysis.get('CC')
    AddPressureProfile = Analysis.get('AddPressureProfile')
    AddPress_z = Analysis.get('AddPress_z')
    AddPress_ez = Analysis.get('DesignAddPress_ez')
    
    Geninf = [format(SlopeFront,'.2f'), 
              format(SlopeBack,'.2f'),
              format(LoadFront,'.2f'), 
              format(LoadBack,'.2f'), 
              format(WaterLevelFront,'.2f'), 
              format(WaterLevelBack,'.2f'), 
              format(WaterDensity,'.2f')]
              
    
    Generalinfo = ""
    
    for item in Geninf: # creates the right amount of space between columns
        if len(str(item)) == 4:
            space = '      ' # 6 spaces
        if len(str(item)) == 5:
            space = '     '  # 5 spaces
        if len(str(item)) == 6:
            space = '    '   # 4 spaces
        Generalinfo += space + str(item)
     
    ############ First lines in SPOOKSWAT input file (L)
    
    L = ['<', 'Project:' +' '+ Project, 'Initials:' +' '+ Initials, 'Subject:' +' '+str(Subject), '>', '<', Generalinfo, '>', '<']
    
    
    
    ################# SOIL FRONT ####################################

    
    for Layer in DesignSoilLayersFront:  # creates lines for soil on front
        
        if State.lower() == 'undrained' and Layer.get('KeepDrained') == 'No':
        
            S_temp = [format(Layer.get('TopLayer'),'.2f'), 
                      format(Layer.get('Gamma_d'),'.2f'), 
                      format(Layer.get('Gamma_m'),'.2f'), 
                      format(Layer.get('i'),'.2f'), 
                      '0.00',
                      format(Layer.get('cu'),'.2f'), 
                      format(Layer.get('r'),'.2f')]
        
        elif State.lower() == 'drained' or Layer.get('KeepDrained') == 'Yes':
            
            S_temp = [format(Layer.get('TopLayer'),'.2f'), 
                      format(Layer.get('Gamma_d'),'.2f'), 
                      format(Layer.get('Gamma_m'),'.2f'), 
                      format(Layer.get('i'),'.2f'),
                      format(Layer.get('phi'),'.2f'), 
                      format(Layer.get('c'),'.2f'), 
                      format(Layer.get('r'),'.2f')]
            
    
        SoilFront = ""
    
        for item in S_temp: # creates the right amount of space between columns
            if len(item) == 4:
                space = '      ' # 6 spaces
            if len(item) == 5:
                space = '     '  # 5 spaces
            if len(item) == 6:
                space = '    '   # 4 spaces
            SoilFront += space + item
        
        L.append(SoilFront)
    
    L.append('>')
    L.append('<')
    
    
    
    
    ########################## SOIL BACK ##############################
    
    for Layer in DesignSoilLayersBack:  # creates lines for soil on front
        
        if State.lower() == 'undrained' and Layer.get('KeepDrained') == 'No':
        
            S_temp = [format(Layer.get('TopLayer'),'.2f'), 
                      format(Layer.get('Gamma_d'),'.2f'), 
                      format(Layer.get('Gamma_m'),'.2f'), 
                      format(Layer.get('i'),'.2f'), 
                      '0.00',
                      format(Layer.get('cu'),'.2f'), 
                      format(Layer.get('r'),'.2f')]
        
        elif State.lower() == 'drained' or Layer.get('KeepDrained') == 'Yes':
            
            S_temp = [format(Layer.get('TopLayer'),'.2f'), 
                      format(Layer.get('Gamma_d'),'.2f'), 
                      format(Layer.get('Gamma_m'),'.2f'), 
                      format(Layer.get('i'),'.2f'),
                      format(Layer.get('phi'),'.2f'), 
                      format(Layer.get('c'),'.2f'), 
                      format(Layer.get('r'),'.2f')]
    
        SoilBack = ""
    
        for item in S_temp: # creates the right amount of space between columns
            if len(item) == 4:
                space = '      ' # 6 spaces
            if len(item) == 5:
                space = '     '  # 5 spaces
            if len(item) == 6:
                space = '    '   # 4 spaces
            SoilBack += space + item
        
        L.append(SoilBack)
    
    L.append('>')
    L.append('<')
            
    
    
    ################## FAILURE MODE (ANCHOR COEFFICIENTS) ###################
    
    
    ## If no anchor
    if AnchorLevel == None: 

        if format(iB,'.2f') != '1.00' or format(iC,'.2f') != '0.00':
            print('No compatibility between failure mode and failure coefficients')
        
            Coeff_temp = ['N/A', 
                          'N/A',
                          'N/A', 
                          'N/A', 
                          'N/A']
        else:
            
            Coeff_temp = [format(iA,'.2f'), 
                          format(iB,'.2f'),
                          format(iC,'.2f'), 
                          format(zT,'.2f'), 
                          format(zR,'.2f')]
        
        Coeff = ""
    
        for item in Coeff_temp: # creates the right amount of space between columns
            if len(item) == 4:
                space = '      ' # 6 spaces
            if len(item) == 5:
                space = '     '  # 5 spaces
            if len(item) == 6:
                space = '    '   # 4 spaces
            Coeff += space + item
            
        Endspace = '                    ' ## 20 end spaces (if no anchor)
        
        L.append(Coeff+Endspace)
        
    # if anchor and no prescribed anchor force
    elif AnchorLevel != None and PrescrbAnchorForce == 0.00: 
    
        Coeff_temp = [format(iA,'.2f'), 
                      format(iB,'.2f'),
                      format(iC,'.2f'), 
                      format(zT,'.2f'), 
                      format(zR,'.2f'),
                      format(AnchorLevel,'.2f')]
    
        Coeff = ""
    
        for item in Coeff_temp: # creates the right amount of space between columns
            if len(item) == 4:
                space = '      ' # 6 spaces
            if len(item) == 5:
                space = '     '  # 5 spaces
            if len(item) == 6:
                space = '    '   # 4 spaces
            Coeff += space + item
            
        Endspace = '          ' ## 10 end spaces (if anchor but no prescribed force)
        
        L.append(Coeff+Endspace)
        
    
    # if anchor and prescribed anchor force
    elif AnchorLevel != None and PrescrbAnchorForce != 0.00 and format(iC,'.2f') == '0.00' and float(iB) > 0: 
    
        Coeff_temp = [format(iA,'.2f'), 
                      format(iB,'.2f'),
                      format(iC,'.2f'), 
                      format(zT,'.2f'), 
                      format(zR,'.2f'),
                      format(AnchorLevel,'.2f'),
                      format(PrescrbAnchorForce,'.2f')]
    
        Coeff = ""
    
        for item in Coeff_temp: # creates the right amount of space between columns
            if len(item) == 4:
                space = '      ' # 6 spaces
            if len(item) == 5:
                space = '     '  # 5 spaces
            if len(item) == 6:
                space = '    '   # 4 spaces
            Coeff += space + item
        
        L.append(Coeff)
    
    
    ############ KING POST WALL PARAMETERS (if data is properly entered in excel input file)
        
    if zB != None and WD != None and CC != None:
        
        KingPost_temp = [format(zB,'.2f'),
                         format(WD,'.2f'),
                         format(CC,'.2f')]
        
        KingPostParam = ""
    
        for item in KingPost_temp: # creates the right amount of space between columns
            if len(str(item)) == 4:
                space = '      ' # 6 spaces
            if len(str(item)) == 5:
                space = '     '  # 5 spaces
            if len(str(item)) == 6:
                space = '    '   # 4 spaces
            KingPostParam += space + str(item)
            
        Initspace = '                                        ' ## 40 inital spaces
        
        L.append(Initspace+KingPostParam)
    
    elif zB == None and WD == None and CC == None:
        
        L.append('')
    
    else:
        
        print('Improper King/Post wall parameters')
        
    
    ###### OUTPUT PATHS (SPOOKS working directory - temporary files)
    
    TemporaryPath = TemporaryWorkingDirectory()
    
    ## output plot file path
    plotpath_output = os.path.join(TemporaryPath, r'spooks.plt')
    
    L.append('>'+'    '+plotpath_output)

    
    ############### ADDITIONAL PRESSURE FILE #########################
    
    if AddPressureProfile in ['AP1', 'AP2', 'AP3', 'AP4', 'AP5', 'AP6', 'AP7', 'AP8', 'AP9', 'AP10']:
    
        ## Additional pressure file
        addpressfile = []
    
        
        APlevel = ""
    
        for item in AddPress_z: # creates the right amount of space between columns

            item = format(float(item),'.2f')
            
            if len(item) == 4:
                space = '      ' # 6 spaces
            if len(item) == 5:
                space = '     '  # 5 spaces
            if len(item) == 6:
                space = '    '   # 4 spaces
            APlevel += space + item
        
        addpressfile.append(APlevel+'    ')
        addpressfile.append('  ')
        
        APvalue = ""
    
        for item in AddPress_ez: # creates the right amount of space between columns
            
            item = format(float(item),'.2f')

            if len(item) == 4:
                space = '      ' # 6 spaces
            if len(item) == 5:
                space = '     '  # 5 spaces
            if len(item) == 6:
                space = '    '   # 4 spaces
            APvalue += space + item
        
        addpressfile.append(APvalue+'    ')
        
        ### Generating additional pressure file for input file
        addpressurefile = os.path.join(TemporaryPath, r'addpressfile')
        
        with open(addpressurefile, 'w+') as f:
            for item in addpressfile:
                f.write("%s\n" % item)
        
        f.close()
                     
        L.append('     '+addpressurefile)
        
    else:
        
        L.append('')
        
    ############### 4.10. END SPOOKSWAT INPUT FILE ###################
    
    L.append('>END')
    
    ###### Export to file
    
    inputfile = os.path.join(TemporaryPath, r'params')
    
    with open(inputfile, 'w+') as f:
        for item in L:
            f.write("%s\n" % item)
    
    f.close()
    
    
    Output = {'Analysis': Analysis,
              'InputFileDir': TemporaryPath,
              'InputFile': inputfile,
              'SPOOKSPlotFile': plotpath_output}
    
    
    return Output
    



def LogFile(InputFileDir,AnalysisNo,SPOOKSOut):
    
    logfile = os.path.join(InputFileDir, r'log_file.txt')
    
    if AnalysisNo == 0:  ### Creates (or overwrites existing) log file
        with open(logfile, 'w') as f:
            for item in SPOOKSOut[13:]:
                f.write("%s\n" % item)
    else:                ### Appends to new log file
         with open(logfile, 'a') as f:
            for item in SPOOKSOut[13:]:
                f.write("%s\n" % item)
    f.close()



def ExecuteSPOOKS(Analysis,logtxt,tk):
    
    Output = GenerateSPOOKSInputFile(Analysis)
    
    Analysis = Output.get('Analysis')
    AnalysisNo = Analysis.get('AnalysisNo')
    InputFile = Output.get('InputFile')
    InputFileDir = Output.get('InputFileDir')
    SPOOKSPlotFile = Output.get('SPOOKSPlotFile')
    
    
    
    ## Date and time
    Now = datetime.now() # current date and time
    DateTime = Now.strftime("%d.%m.%Y, %H:%M:%S")
    Date = Now.strftime("%d.%m.%Y")
    
    args = r'spookswat.exe'+' '+'/CALC:"'+InputFile+'"'
        
    process = subprocess.Popen(args, cwd = r'C:\Program Files (x86)\WinSpooks', shell=True, stdin = None, stderr = subprocess.PIPE, stdout = subprocess.PIPE, universal_newlines=True)
    
    SPOOKSOut = process.stdout.readlines()
    
    logtxt.insert(tk.END, SPOOKSOut)
    
    print(SPOOKSOut) ### Printing output if script run as script

    #logtxt.insert(tk.END, out) ### Print "out" to log tab in GUI
    
    ReportWarnings = []
    ReportErrors = []
    
    ###### If "STOPPED", "WARNING" or "*ERR*" is contained in SPOOKSWAT output, status changes
    for i in range(0,len(SPOOKSOut)):
        p = str(SPOOKSOut[i])
        if 'STOPPED' in p:
            #stat.configure(text = 'Calculation failed - check log file')
            print('Calculation stopped - check log file')
            ReportErrors.append('Calculation stopped - check log file')
            #break_calc == 'break'
        if 'WARNING' in p:
            #warning.configure(text = out[i])
            #stat.configure(text = 'Calculation stopped')
            print('Calculation stopped')
            #break_calc == 'break'
            ReportWarnings.append('WinSPOOKS warning: check log file')
        if '*ERR*' in p:
            #warning.configure(text = out[i])
            #stat.configure(text = 'Calculation stopped')
            #break_calc == 'break'
            print('Calculation stopped due to error')
            ReportErrors.append('WinSPOOKS error: check log file')
        if 'The input contains boundaries below the encastre level.' in p: ## warning in report if generation of reports is checked off:
            ReportWarnings.append('The input contains boundaries below the encastre level.')
            
    LogFile(InputFileDir,AnalysisNo,SPOOKSOut)
    
    ExecuteOutput = {'Analysis': Analysis,
                     'Date': Date,
                     'DateTime': DateTime,
                     'SPOOKSOutput': SPOOKSOut,
                     'Warnings': ReportWarnings,
                     'Errors': ReportErrors,
                     'InputFileDir': InputFileDir,
                     'SPOOKSPlotFile': SPOOKSPlotFile}
    
    
    return ExecuteOutput
    


def GetResults(ExecuteOutput):
    
    Analysis = ExecuteOutput.get('Analysis')
    SPOOKSOutput = ExecuteOutput.get('SPOOKSOutput')
    Warnings = ExecuteOutput.get('Warnings')
    Errors = ExecuteOutput.get('Errors')
    InputFileDir = ExecuteOutput.get('InputFileDir')
    SPOOKSPlotFile = ExecuteOutput.get('SPOOKSPlotFile')
    Date = ExecuteOutput.get('Date')
    DateTime = ExecuteOutput.get('DateTime')
    
    
    ## Initialise variables
    encastrelvl = None
    encastremoment = None
    yieldhingelvl = None
    yieldhingemoment = None
    toelvl = None
    anchforc = None
    anchmom = None
    

    for Line in SPOOKSOutput:
        
        if 'ENCASTRE LEVEL' in Line:
            encastrelvl = Line.split()
            encastrelvl = float(encastrelvl[-1])
        elif encastrelvl == None:
            encastrelvl = 'N/A'
            
        if 'ENCASTRE MOMENT' in Line:
            encastremoment = Line.split()
            encastremoment = float(encastremoment[-2])
        elif encastremoment == None:
            encastremoment = 'N/A'
            
        if 'LEVEL OF YIELD HINGE' in Line:
            yieldhingelvl = Line.split()
            yieldhingelvl = float(yieldhingelvl[-1])
        elif yieldhingelvl == None:
            yieldhingelvl = 'N/A'
            
        if 'MOMENT IN YIELD HINGE' in Line:
            yieldhingemoment = Line.split()
            yieldhingemoment = float(yieldhingemoment[-2])
        elif yieldhingemoment == None:
            yieldhingemoment = 'N/A'
            
        if 'FOOT LEVEL' in Line or 'LEVEL OF FOOT' in Line:
            toelvl = Line.split()
            toelvl = float(toelvl[-1])
        elif toelvl == None:
            toelvl = 'N/A'
            
        if 'ANCHOR FORCE' in Line:
            anchforc = Line.split()
            anchforc = float(anchforc[-4])
        elif anchforc == None:
            anchforc = 'N/A'
            
        if 'MOMENT AT ANCHOR' in Line:
            anchmom = Line.split()
            anchmom = float(anchmom[-2])
        elif anchmom == None:
            anchmom = 'N/A'
    
        
    
    ############# Extracting earth pressure results ###############
        
    results = open(SPOOKSPlotFile,'r').readlines()
    
    # Finds index of "Kote  e1  e2  dw e-net   T   M ju"
    index_header = [i for i, elem in enumerate(results) if 'Kote' in elem]
    index_header = int(index_header[0]) # converting index to integer (for slicing "results")
    
    ### Finds empty strings
    index_empty = []
    for counter, value in enumerate(results):
        if value.isspace() == True:
            index_empty = np.append(index_empty,counter)
            
    # Empty string closest to and greater than header (index-wise)
           # Removing index values less than header
    index_empty = list(filter(lambda a: a > index_header, index_empty))
           # Lower boundary of earth pressure results
    boundary_low = int(index_empty[0]) # converting index to integer (for slicing "results")
    
    # Earth pressure results (slicing)
    results = results[index_header:boundary_low]
    # Split strings (parse results into columns)
    EarthPressureResults = np.empty((0,8), int) # Empty array
    
    
    level_graphic = [] ## level column
    e1_graphic = []    ## earth pressure on front side
    e2_graphic = []    ## earth pressure on back side
    dw_graphic = []    ## differential hydrostatic pressure
    enet_graphic = []  ## net earth pressure
    shearforce = []
    shearlevel = []
    moment = []
    momentlevel = []
    ju_graphic = []
    
    for x in results:
        
        splitstrings = np.array(x.split())
        splitstrings = splitstrings.reshape((1,-1)) # transpose
        EarthPressureResults = np.append(EarthPressureResults, splitstrings, axis = 0) 
        
        ## Splitting columns from results
        if 'T' not in x: ## Removes "T" (header) from shear force column and "M" from moment column
            level_graphic.append(float(splitstrings[:,0]))
            e1_graphic.append(float(splitstrings[:,1])*-1) ### *-1 is added for graphical representation
            e2_graphic.append(float(splitstrings[:,2]))
            dw_graphic.append(float(splitstrings[:,3]))
            enet_graphic.append(float(splitstrings[:,4]))
            shearforce.append(float(splitstrings[:,5]))
            shearlevel.append(float(splitstrings[:,0]))
            moment.append(float(splitstrings[:,6]))
            momentlevel.append(float(splitstrings[:,0]))
            ju_graphic.append(float(splitstrings[:,7]))
            
    
    ## Max. shear force from absolute values
    maxshear = max([abs(val) for val in shearforce])
    index = [i for i, elem in enumerate(shearforce) if maxshear == abs(elem)]
    index = index[0]
    maxshearlevel = shearlevel[index]
    maxshear = shearforce[index]
    ## Max. moment from absolute values
    maxmom = max([abs(val) for val in moment])
    index = [i for i, elem in enumerate(moment) if maxmom == abs(elem)]
    index = index[0]
    maxmomlvl = momentlevel[index]
    maxmom = moment[index]

    
    PlotResults = {'MaxShearForce': maxshear,
                   'LevelMaxShearforce': maxshearlevel,
                   'MaxMoment': maxmom,
                   'LevelMaxMoment': maxmomlvl,
                   'PlotLevels': level_graphic,
                   'e1': e1_graphic,
                   'e2': e2_graphic,
                   'DW': dw_graphic,
                   'enet': enet_graphic,
                   'ShearForce': shearforce,
                   'ShearForceLevel': shearlevel,
                   'Moment': moment,
                   'MomentLevel': momentlevel,
                   'JU': ju_graphic}
        
        
    Results = {'EncastreLevel': encastrelvl,
               'EncastreMoment': encastremoment,
               'YieldHingeLevel': yieldhingelvl,
               'YieldHingeMoment': yieldhingemoment,
               'ToeLevel': toelvl,
               'AnchorForce': anchforc,
               'MomentAtAnchor': anchmom}
    
    
    GetResultsOutput = {'Analysis': Analysis,
                        'DateTime': DateTime,
                        'Date': Date,
                        'Warnings': Warnings,
                        'Errors': Errors,
                        'PlotResults': PlotResults,
                        'EarthPressureResults': EarthPressureResults,
                        'Results': Results}
    
    return GetResultsOutput




def SPOOKSFeeder(input_path,calcno,pb,tabcalc,logtxt,tk):
    
    
    FeederOutput = []
    
    
    GeneratedAnalyses = GenerateAnalyses(input_path)
    
    
    ## Progress bar maximum
    pb['maximum'] = len(GeneratedAnalyses)-1
    pb.update() ## update progress bar
    
    ## Loop through all generated analyses and append output to FeederOutput
    ReportErrors = []
    for Analysis in GeneratedAnalyses:
        #print(Analysis)
        
        ### PROGRESS BAR AND CALCULATION NO UPDATE
        AnalysisNo = Analysis.get('AnalysisNo')
        calcno.configure(text = str(AnalysisNo)) ## Calculation number for GUI
        tabcalc.update_idletasks()
        pb['value'] = AnalysisNo
        pb.update() ## update progress bar
        
        ExecuteOutput = ExecuteSPOOKS(Analysis,logtxt,tk)
        
        ReportErrors = ExecuteOutput.get('Errors')
    
        GetResultsOutput = GetResults(ExecuteOutput)
    
    
    #else:
    #    print("FAILED")
    #    calcno.configure(text = 'Calculation failed') ## Calculation number for GUI
    #    tabcalc.update_idletasks()
    #    break

            
        
        FeederOutput.append({'FeedAnalysis': Analysis,
                             'ExecuteOutput': ExecuteOutput,
                             'GetResultsOutput': GetResultsOutput})
        
    
    return FeederOutput
        

def VerticalEquilibrium(GetResultsOutput):
    
    ### tangential earthpressure calculated as qs = e * tan(delta) + adhesion , where r = tan(delta)/tan(phi) = adhesion/cohesion
    ### for undrained layers qs = cu * r
    
    Analysis = GetResultsOutput.get('Analysis')
    AxialWallLoad = Analysis.get('AxialWallLoad')
    State = Analysis.get('State')
    DesignSoilLayersFront = Analysis.get('DesignSoilLayersFront')
    DesignSoilLayersBack = Analysis.get('DesignSoilLayersBack')
    PlotResults = GetResultsOutput.get('PlotResults')
    e1 = PlotResults.get('e1')
    e2 = PlotResults.get('e2')
    PlotLevels = PlotResults.get('PlotLevels')
    
    ### Max soil layer level back
    SoilLevelsBack = []
    for SoilLayer in DesignSoilLayersBack:
        SoilLevelsBack.append(SoilLayer.get('TopLayer'))
    GroundLevelBack = max(SoilLevelsBack)
    
    ### Max soil layer level front
    SoilLevelsFront = []
    for SoilLayer in DesignSoilLayersFront:
        SoilLevelsFront.append(SoilLayer.get('TopLayer'))
    GroundLevelFront = max(SoilLevelsFront)
    
    
    ## Initialise
    sum_tanforce_back = 0  
    sum_tanforce_front = 0
    
    
    for i in range(len(e1)-1):
    
        ##### back side of wall (active)
        
        tan_force_back = 0
        
        if PlotLevels[i+1] < GroundLevelBack: ## makes sure only earth pressures lower than ground level on back are considered.
        
            force_back = abs(np.trapz(e2[i:i+2], PlotLevels[i:i+2]))  ## sums area under earth pressure curve

            ## finds index of nearest soil stratigraphy (horizon) above earth pressure area
            index = [j for j, k in enumerate(SoilLevelsBack) if float(k) > PlotLevels[i+1]]
            index = index[-1]
            
            SoilLayer = DesignSoilLayersBack[index]

            
            if State.lower() == 'drained' or SoilLayer.get('KeepDrained') == 'Yes':
                
                tan_delta_back = float(SoilLayer.get('r')) * np.tan(np.radians(SoilLayer.get('phi'))) ## wall friction angle
                
                ## Tangential force back
                tan_force_back = force_back * tan_delta_back + float(SoilLayer.get('c')) * (PlotLevels[i]-PlotLevels[i+1]) * float(SoilLayer.get('r'))
                
            
            elif State.lower() == 'undrained' and SoilLayer.get('KeepDrained') == 'No':
                
                tan_force_back = float(SoilLayer.get('cu')) * (PlotLevels[i]-PlotLevels[i+1]) * float(SoilLayer.get('r'))
        
        ## Front side of wall (passive)
            
        tan_force_front = 0
        
        if PlotLevels[i+1] < GroundLevelFront:
        
            force_front = abs(np.trapz(e1[i:i+2], PlotLevels[i:i+2])) ## sums area under earth pressure curve

            if PlotLevels[i] == GroundLevelFront : ## finds index of nearest soil stratigraphy (horizon) above earth pressure area
                index = 0 
            else:
                index = [j for j, k in enumerate(SoilLevelsFront) if float(k) > PlotLevels[i+1]]
                index = index[-1]
            
            SoilLayer = DesignSoilLayersFront[index]
            
            if State.lower() == 'drained' or SoilLayer.get('KeepDrained') == 'Yes':
              
                tan_delta_front = float(SoilLayer.get('r')) * np.tan(np.radians(SoilLayer.get('phi'))) ## wall friction angle
            
                ## Tangential force front
                tan_force_front = force_front * tan_delta_front+ float(SoilLayer.get('c')) * (PlotLevels[i]-PlotLevels[i+1]) * float(SoilLayer.get('r'))
            
            
            elif State.lower() == 'undrained' and SoilLayer.get('KeepDrained') == 'No':
                
                tan_force_front = float(SoilLayer.get('cu')) * (PlotLevels[i]-PlotLevels[i+1]) * float(SoilLayer.get('r'))
                
        
        else:
        
            tan_force_front = 0
            
        ## Sum of tangential forces back    
        sum_tanforce_back = sum_tanforce_back + abs(tan_force_back)
        ## Sum of tangential forces front
        sum_tanforce_front = sum_tanforce_front + abs(tan_force_front)
    
    ## Total sum of tangential forces
    sum_tanforce = sum_tanforce_front - sum_tanforce_back - AxialWallLoad
    
    
    VerticalEquilibriumOutput = {'Analysis': Analysis,
                                 'GetResultsOutput': GetResultsOutput,
                                 'SumTanForceBack':  sum_tanforce_back,
                                 'SumTanForceFront': sum_tanforce_front,
                                 'SumTanForce':      sum_tanforce}
        
        
    return VerticalEquilibriumOutput


def ReportFront(VerticalEquilibriumOutput,OutputDirList,Version):
    
    print('Generating report front...')

    GetResultsOutput = VerticalEquilibriumOutput.get('GetResultsOutput')
        
    Analysis = VerticalEquilibriumOutput.get('Analysis')
    Project = Analysis.get('Project')
    Subject = Analysis.get('Subject')
    AnalysisNo = Analysis.get('AnalysisNo')
    State = Analysis.get('State')
    SoilLayersFront = Analysis.get('SoilLayersFront')
    SoilLayersBack = Analysis.get('SoilLayersBack')
    PlotResults = GetResultsOutput.get('PlotResults')
    PlotLevels = PlotResults.get('PlotLevels')
    e1 = PlotResults.get('e1')
    e2 = PlotResults.get('e2')
    Moment = PlotResults.get('Moment')
    MaxMoment = PlotResults.get('MaxMoment')
    ShearForce = PlotResults.get('ShearForce')
    zT = Analysis.get('zT')
    SlopeFront = Analysis.get('SlopeFront')
    SlopeBack = Analysis.get('SlopeBack')
    LoadFront = Analysis.get('LoadFront')
    LoadBack = Analysis.get('LoadBack')
    zR = Analysis.get('LevelLoadBack')
    WaterLevelFront = Analysis.get('WaterLevelFront')
    WaterLevelBack = Analysis.get('WaterLevelBack')
    AnchorLevel = Analysis.get('AnchorLevel')
    AnchorForce = GetResultsOutput.get('Results').get('AnchorForce')
    AnchorInclination = Analysis.get('AnchorInclination')
    SumTanForce = VerticalEquilibriumOutput.get('SumTanForce')
    WallSpecificWeight = Analysis.get('WallMass') * 9.82 / 1000
    DateTime = GetResultsOutput.get('DateTime')
    OutputDir = OutputDirList[-1]
    
    ### Max soil layer level back
    SoilLevelsBack = []
    for SoilLayer in SoilLayersBack:
        SoilLevelsBack.append(SoilLayer.get('TopLayer'))
    GroundLevelBack = max(SoilLevelsBack)
    
    ### Max soil layer level front
    SoilLevelsFront = []
    for SoilLayer in SoilLayersFront:
        SoilLevelsFront.append(SoilLayer.get('TopLayer'))
    GroundLevelFront = max(SoilLevelsFront)
    
    ToeLevel = GetResultsOutput.get('Results').get('ToeLevel')
    if ToeLevel == 'N/A':
        ToeLevel = GroundLevelFront - 0.1
    print(ToeLevel)
    
    try:
        WeightWallTotal = WallSpecificWeight * (zT - ToeLevel)
    except TypeError:
        WeightWallTotal = 'N/A'
    print(WeightWallTotal)
    
    
    
    
    ### Turning interactive plotting off
    plt.ioff()
        
    ## Customising math text
    plt.rcParams.update(plt.rcParamsDefault)
    
    plt.rcParams['mathtext.fontset'] = 'custom'
    plt.rcParams['mathtext.rm'] = 'Arial'
    
    ######## Cubic interpolation of moment curve (smoothing of moment curve)##
    level_plot_unique = [] ## list of unique level values
    m_plot_unique = []
    for i in range(len(PlotLevels)-1):
        if PlotLevels[i] != PlotLevels[i+1]:
            level_plot_unique.append(PlotLevels[i])
            m_plot_unique.append(Moment[i])
    
    level_plot_unique.append(PlotLevels[-1])
    m_plot_unique.append(Moment[-1])
    
    x = np.asarray(level_plot_unique)
    y = np.asarray(m_plot_unique)
    m_curve = interp1d(x, y, kind='quadratic')
    
    curve_level = np.linspace(min(level_plot_unique),max(level_plot_unique) , 100)
    ############## End of cubic interpolation ###############################
    

    
    #### Ground levels for graphic representation
    maxval = max(max(ShearForce),max(e1),max(e2),max(Moment))
    minval = min(min(ShearForce),min(e1),min(e2),min(Moment))
    grnextent = max(abs(minval),abs(maxval))
    ## Plot limits
    xmin = grnextent*1.05*-1
    xmax = grnextent*1.05
    ymin = ToeLevel-0.02*(float(zT)-ToeLevel)
    ymax = float(zT)+0.1*(float(zT)-ToeLevel)
    ### Aspect ratio (for graphical proportionality)
    aspectratio = ((xmax-xmin)/6.41) / ((ymax-ymin)/9.00)  ## 6.41 and 9.00 are the respective frame dimensions
    grnextent_front = [0, grnextent*-1]
    grnlevel_front = [GroundLevelFront, GroundLevelFront-grnextent_front[1]*np.tan(np.radians(SlopeFront))/aspectratio]
    grnextent_back = [0,grnextent]
    grnlevel_back = [GroundLevelBack, GroundLevelBack+grnextent_back[1]*np.tan(np.radians(SlopeBack))/aspectratio]
    
    #### Water levels for graphic representation
    waterlvl_front = [WaterLevelFront-0.005*(ymax-ymin), WaterLevelFront-0.005*(ymax-ymin)]
    waterlvl_back = [WaterLevelBack-0.005*(ymax-ymin), WaterLevelBack-0.005*(ymax-ymin)]
        
    #### Graphic wall parameters
    wall_x = [0,0] ## horisontal location of wall
    wall_y = [zT, ToeLevel]

    ### Assigning figure for report
    fig = plt.figure(num=None, figsize=(8, 8), dpi=80, facecolor='w', edgecolor='k')
    
    ### Forces, pressures and structures
    ## Earth pressures
    plt.plot(e1,PlotLevels, alpha = 1.0, color = 'mediumseagreen', label = '$e_{1}$, $e_{2}$ [kPa]', linewidth = 1.0)
    plt.plot(e2,PlotLevels, alpha = 1.0, color = 'mediumseagreen', linewidth = 1.0)
    ## Structural forces
    ### Smoothened moment curve
    plt.plot(m_curve(curve_level),curve_level, alpha = 1.0, color = 'red', label = '$M_{Ed}$ [kNm/m]', linewidth = 1.0)
    plt.plot(ShearForce,PlotLevels, alpha = 1.0, color = 'grey', label = '$V_{Ed}$ [kN/m]', linewidth = 1.0)
    ## Sheet pile wall
    plt.plot(wall_x,wall_y, alpha = 1.0, color = 'black', linewidth = 1.5)
    ## Ground level front/back
    plt.plot(grnextent_back,grnlevel_back, alpha = 1.0, color = 'black', linewidth = 1.5)
    plt.plot(grnextent_front,grnlevel_front, alpha = 1.0, color = 'black', linewidth = 1.5)
    ## Water level fron/back
    plt.plot(grnextent_back, waterlvl_back, alpha = 1.0, color = 'dodgerblue', linewidth = 1.2, linestyle='dashed')
    plt.plot(grnextent_front, waterlvl_front, alpha = 1.0, color = 'dodgerblue', linewidth = 1.2, linestyle='dashed')
    
    ### Anchor
    if AnchorLevel != None:
        ## Axial anchor force
        AnchorAxial = AnchorForce/np.cos(np.radians(AnchorInclination))
        plt.annotate("", xy=(1.5*aspectratio, AnchorLevel), xytext=(0.0, AnchorLevel), arrowprops=dict(headwidth=15, headlength=10, width=0.1, color='black'))

    
    ### WinSpooks Plug-in text
    props = dict(boxstyle='round', facecolor='white', alpha=0.0)
    plt.text(grnextent*-1, ymin+(ymax-ymin)*0.02, 'COWI WinSpooks Plug-in '+Version+' ,'+DateTime+'\n'+'Saved to: '+OutputDir, fontsize=5.0, verticalalignment='top', bbox=props)
    
    plt.ylabel('Level [m]')
    plt.title(Project+', '+Subject+', '+'Calc. no. '+str(AnalysisNo))
    plt.ylim(top=ymax, bottom=ymin)
    plt.xlim(xmin, xmax)
    plt.legend(loc='upper right')
    
    ### Text box soil parameters
    for i in range(0,len(SoilLayersBack)):  ## soil stratigraphies back excluding ground level
        
        SoilLayer = SoilLayersBack[i]
        
        if float(SoilLayer.get('TopLayer')) > ToeLevel: ## does not plot soil layers deeper than toe level
        
            if i > 0:
                Stratigraphy = [SoilLayer.get('TopLayer'), SoilLayer.get('TopLayer')]
                plt.plot(grnextent_back, Stratigraphy, alpha = 0.8, color = 'black', linewidth = 1.0)
            
            if State.lower() == 'drained' or SoilLayer.get('KeepDrained') == 'Yes':
                                        
                textstr = ', '.join((
                r'$\rm \gamma_d/ \gamma_m=%.0f/%.0f$ $\rm kN/m^{3}$' % (SoilLayer.get('Gamma_d'),SoilLayer.get('Gamma_m')),
                r'$\rm \phi\prime_{k}=%.0f ^{\circ}$' % (SoilLayer.get('phi'), ),
                r'$\rm c\prime_{k}=%.0f$ $\rm kPa$' % (SoilLayer.get('c'), ),
                r'$\rm r=%.1f$' % (SoilLayer.get('r'), ),
                r'$\rm i=%.1f$' % (SoilLayer.get('i'), )))
                textstr = '\n'.join((SoilLayer.get('Description'),textstr))
                props = dict(boxstyle='round', facecolor='grey', alpha=0.0)
            
            elif State.lower() == 'undrained' and SoilLayer.get('KeepDrained') == 'No':
                
                textstr = ', '.join((
                r'$\rm \gamma_d/\gamma_m=%.0f/%.0f$ $\rm kN/m^{3}$' % (SoilLayer.get('Gamma_d'),SoilLayer.get('Gamma_m')),
                r'$\rm cu_{k}=%.0f$ $\rm kPa$' % (SoilLayer.get('cu'), ),
                r'$\rm r=%.1f$' % (SoilLayer.get('r'), ),
                r'$\rm i=%.1f$' % (SoilLayer.get('i'), )))
                textstr = '\n'.join((SoilLayer.get('Description'),textstr))
                props = dict(boxstyle='round', facecolor='grey', alpha=0.0)
                
            
            if i != len(SoilLayersBack)-1 and SoilLayersBack[i+1].get('TopLayer') > ToeLevel: ## Vertical position of soil parameters data
                VertPos = (SoilLayer.get('TopLayer') + SoilLayersBack[i+1].get('TopLayer'))/2
            else:
                VertPos = (SoilLayer.get('TopLayer') + ToeLevel) / 2
            plt.text(grnextent*0.05, VertPos, textstr, fontsize=8.0, verticalalignment='center', bbox=props)
    
    
    for i in range(0,len(SoilLayersFront)):  ## soil stratigraphies front excluding ground level
        
        SoilLayer = SoilLayersFront[i]
        
        if float(SoilLayer.get('TopLayer')) > ToeLevel: ## does not plot soil layers deeper than toe level
        
            if i > 0:
                Stratigraphy = [SoilLayer.get('TopLayer'), SoilLayer.get('TopLayer')]
                plt.plot(grnextent_front, Stratigraphy, alpha = 0.8, color = 'black', linewidth = 1.0)
            
            if State.lower() == 'drained' or SoilLayer.get('KeepDrained') == 'Yes':
                
                textstr = ', '.join((
                r'$\rm \gamma_d/ \gamma_m=%.0f/%.0f$ $\rm kN/m^{3}$' % (SoilLayer.get('Gamma_d'),SoilLayer.get('Gamma_m')),
                r'$\rm \phi\prime_{k}=%.0f ^{\circ}$' % (SoilLayer.get('phi'), ),
                r'$\rm c\prime_{k}=%.0f$ $\rm kPa$' % (SoilLayer.get('c'), ),
                r'$\rm r=%.1f$' % (SoilLayer.get('r'), ),
                r'$\rm i=%.1f$' % (SoilLayer.get('i'), )))
                textstr = '\n'.join((SoilLayer.get('Description'),textstr))
                props = dict(boxstyle='round', facecolor='grey', alpha=0.0)
            
            elif State.lower() == 'undrained' and SoilLayer.get('KeepDrained') == 'No':
                
                textstr = ', '.join((
                r'$\rm \gamma_d/\gamma_m=%.0f/%.0f$ $\rm kN/m^{3}$' % (SoilLayer.get('Gamma_d'),SoilLayer.get('Gamma_m')),
                r'$\rm cu_{k}=%.0f$ $\rm kPa$' % (SoilLayer.get('cu'), ),
                r'$\rm r=%.1f$' % (SoilLayer.get('r'), ),
                r'$\rm i=%.1f$' % (SoilLayer.get('i'), )))
                textstr = '\n'.join((SoilLayer.get('Description'),textstr))
                props = dict(boxstyle='round', facecolor='grey', alpha=0.0)
                
            
            if i != len(SoilLayersFront)-1 and SoilLayersFront[i+1].get('TopLayer') > ToeLevel: ## Vertical position of soil parameters data
                VertPos = (SoilLayer.get('TopLayer') + SoilLayersFront[i+1].get('TopLayer'))/2
            else:
                VertPos = (SoilLayer.get('TopLayer') + ToeLevel) / 2
            plt.text(grnextent_front[1], VertPos, textstr, fontsize=8.0, verticalalignment='center', bbox=props)
            
            
    ### PLOT LOADS
    ## front load
    textstr = ''.join(r'$\rm q_{fk}=%.0f$ $\rm kPa$' % (LoadFront), )
    plt.text((grnextent_front[0] + grnextent_front[1])/2, grnlevel_front[0], textstr, fontsize=8.0, verticalalignment='bottom', bbox=props)
    ## back load
    textstr = ''.join(r'$\rm q_{bk}=%.0f$ $\rm kPa$' % (LoadBack), )
    plt.text((grnextent_back[0] + grnextent_back[1])/2, grnlevel_back[0], textstr, fontsize=8.0, verticalalignment='bottom', bbox=props)
        

    ### Text box results
    if AnchorLevel != None:
        textstr = '\n'.join((
        r'$\rm |M_{Ed,max}|=%.0f$ kNm/m' % (float(abs(MaxMoment)), ),
        #r'$\rm V_{Ed,max}=%.1f$ kN/m' % (vmax, ),
        r'$\rm Toe_{level}=%.1f$ m' % (float(ToeLevel), ),
        r'$\rm A_{d}=%.0f$ kN/m' % (float(AnchorForce), ),
        r'$\rm \Sigma F↑=%.0f$ kN/m' % (float(SumTanForce-WeightWallTotal-AnchorAxial*np.sin(np.radians(AnchorInclination))), )))
        props = dict(boxstyle='round', facecolor='grey', alpha=0.1)
        plt.text(-1*grnextent, ymax-0.017*(ymax-ymin), textstr, fontsize=10, verticalalignment='top', bbox=props)
   
    else:
        textstr = '\n'.join((
        r'$\rm |M_{Ed,max}|=%.0f$ kNm/m' % (float(abs(MaxMoment)), ),
        #r'$\rm V_{Ed,max}=%.1f$ kN/m' % (vmax, ),
        r'$\rm Toe_{level}=%.1f$ m' % (float(ToeLevel), ),
        r'$\rm \Sigma F↑=%.0f$ kN/m' % (float(SumTanForce-WeightWallTotal), )))
        props = dict(boxstyle='round', facecolor='grey', alpha=0.1)
        plt.text(-1*grnextent, ymax-0.017*(ymax-ymin), textstr, fontsize=10, verticalalignment='top', bbox=props)
   
    
    ### Text box logo
    plt.text(grnextent, ymin,'COWI', fontsize=20, fontname='Century Schoolbook', color='orangered',ha='right', va='bottom', alpha=1.0)
    
    fig.set_size_inches(8.27,11.69)

    
    ## Save report front plot in temporary working directory
    
    TemporaryPath = TemporaryWorkingDirectory()
    
    TempReportFrontPath = os.path.join(TemporaryPath, r'ReportFront.pdf')
    
    # plt.savefig(TempReportFrontPath, dpi='figure', facecolor='w', edgecolor='w',
    #     orientation='portrait', format='pdf',
    #     transparent=False, bbox_inches=None, pad_inches=0.05,
    #     frameon=None, metadata=None)
    
    plt.savefig(TempReportFrontPath, dpi='figure', facecolor='w', edgecolor='w',
        orientation='portrait', format='pdf',
        transparent=False, bbox_inches=None, pad_inches=0.05, metadata=None)
    
    
    ## Close figures
    plt.close()
    
    return TempReportFrontPath


def PDFGenerator(VerticalEquilibriumOutput, SheetPileAddOnResults, Version):
    
    print('Generating report pages...')

    
    GetResultsOutput = VerticalEquilibriumOutput.get('GetResultsOutput')
    
        
    Analysis = VerticalEquilibriumOutput.get('Analysis')
    Project = Analysis.get('Project')
    Subject = Analysis.get('Subject')
    Initials = Analysis.get('Initials')
    Checker = Analysis.get('Checker')
    Approver = Analysis.get('Approver')
    Date = GetResultsOutput.get('Date')
    Warnings = GetResultsOutput.get('Warnings')
    AnalysisNo = Analysis.get('AnalysisNo')
    State = Analysis.get('State')
    WaterDensity = Analysis.get('WaterDensity')
    SoilProfile = Analysis.get('SoilProfile')
    SoilLayersFront = Analysis.get('SoilLayersFront')
    SoilLayersBack = Analysis.get('SoilLayersBack')
    WaterLevelFront = Analysis.get('WaterLevelFront')
    WaterLevelBack = Analysis.get('WaterLevelBack')
    AddPressureProfile = Analysis.get('AddPressureProfile')
    AddPress_z = Analysis.get('AddPress_z')
    AddPress_ez = Analysis.get('AddPress_ez')
    AddPress_ez_Design = Analysis.get('DesignAddPress_ez')
    SlopeFront = Analysis.get('SlopeFront')
    SlopeBack = Analysis.get('SlopeBack')
    LoadFront = Analysis.get('LoadFront')
    LoadBack = Analysis.get('LoadBack')
    AxialWallLoad = Analysis.get('AxialWallLoad')
    zR = Analysis.get('LevelLoadBack')
    Alpha = Analysis.get('Alpha')
    ConsequenceClass = Analysis.get('ConsequenceClass')
    PartialSafetyFactors = Analysis.get('PartialSafetyFactors')
    AnchorLevel = Analysis.get('AnchorLevel')
    AnchorForce = GetResultsOutput.get('Results').get('AnchorForce')
    AnchorInclination = Analysis.get('AnchorInclination')
    PrescrbAnchorForce = Analysis.get('PrescrbAnchorForce')
    iA = Analysis.get('iA')
    iB = Analysis.get('iB')
    iC = Analysis.get('iC')
    zB = Analysis.get('zB')
    WD = Analysis.get('WD')
    CC = Analysis.get('CC')
    PlotResults = GetResultsOutput.get('PlotResults')
    MaxMoment = PlotResults.get('MaxMoment')
    MaxShearForce = PlotResults.get('MaxShearForce')
    MomentAtAnchor = GetResultsOutput.get('Results').get('MomentAtAnchor')
    zT = Analysis.get('zT')
    ToeLevel = GetResultsOutput.get('Results').get('ToeLevel')
    ### Max soil layer level front
    
    ToeLevel = GetResultsOutput.get('Results').get('ToeLevel')
    if ToeLevel == 'N/A':
        SoilLevelsFront = []
        for SoilLayer in SoilLayersFront:
            SoilLevelsFront.append(SoilLayer.get('TopLayer'))
        GroundLevelFront = max(SoilLevelsFront)
        ToeLevel = GroundLevelFront - 0.1
    print(ToeLevel)
    SumTanForce = VerticalEquilibriumOutput.get('SumTanForce')
    WallMass = Analysis.get('WallMass')
    WallSpecificWeight = Analysis.get('WallMass') * 9.82 / 1000
    WeightWallTotal = WallSpecificWeight * (zT - ToeLevel)

    PlotLevels = PlotResults.get('PlotLevels')
    e1 = PlotResults.get('e1')
    e2 = PlotResults.get('e2')
    Moment = PlotResults.get('Moment')
    ShearForce = PlotResults.get('ShearForce')
    DW = PlotResults.get('DW')
    ENet = PlotResults.get('enet')
    JU = PlotResults.get('JU')
    
    
    
    ##################### Sheet pile add on #######################
    
    SheetPileAddOnInput = Analysis.get('SheetPileAddOnInput')
    

    
    UseAddOn = SheetPileAddOnInput.get('UseAddOn')
    LimitState = SheetPileAddOnInput.get('LimitState')
    ControlClass = SheetPileAddOnInput.get('ControlClass')
    Optimize = SheetPileAddOnInput.get('Optimize')
    MaxUtilization = SheetPileAddOnInput.get('MaxUtilization')
    fyk = SheetPileAddOnInput.get('fyk')
    BetaB = SheetPileAddOnInput.get('BetaB')
    BetaD = SheetPileAddOnInput.get('BetaD')
    DesignLife = SheetPileAddOnInput.get('DesignLife')
    tCor = SheetPileAddOnInput.get('tCor')
    tCorLevel = SheetPileAddOnInput.get('tCorLevel')
    SoilDeposit = SheetPileAddOnInput.get('SoilDeposit')
    
    ### results
    Sheetpile = SheetPileAddOnResults.get('SheetPileProfile')
    u_rel = SheetPileAddOnResults.get('RUR')
    control_Rot = SheetPileAddOnResults.get('RotCap')
    u_rel_lvl = SheetPileAddOnResults.get('RURLevel')

    


    
        
    # ## Font for cowi logo
    # pdf.add_font('century', '', r"C:\Users\EMBT\OneDrive - COWI\Documents\Python\SPOOKS\CENSCBK.TTF", uni=True)
    
    class PDF(FPDF):
        def header(self):
            self.set_font('Arial', '', 20)
            self.cell(0, 10, 'COWI', 0, 0, 'L')
            self.set_font('Courier', '', 9)
            self.cell(0, 10, Project+', '+Subject+', '+Date, 0, 0, 'R')
            # Line break
            self.ln(10)
    
        # Page footer
        def footer(self):
            # Position at 1.5 cm from bottom
            self.set_y(-15)
            self.set_font('Courier', '', 7)
            self.cell(0, 10, 'COWI WinSpooks Plug-in '+Version, 0, 0, 'L')
            # Page number
            self.set_font('Courier', '', 9)
            self.cell(0, 10, 'Page ' + str(self.page_no()) + ' / {nb}', 0, 0, 'R')
            
    
    # Instantiation of inherited class
    pdf = PDF()
    pdf.alias_nb_pages()
    
    # Add a page 
    pdf.add_page('P') 
      
    # style and size of font  
    pdf.set_font("Courier", size = 12) 
    
    # Effective page width
    epw = pdf.w - 2*pdf.l_margin
    ## Width of columns
    col_width1 = epw/3
    col_width2 = epw/2
    # Text height is the same as current font size
    th = pdf.font_size
    
    #### WARNINGS
    pdf.set_font("Courier", size = 15) 
    pdf.cell(200, 10, txt = '0. Warnings', 
              ln = 6, align = 'L')
    pdf.ln(2)
    ## Warning
    pdf.set_font("Courier", size = 12)
    if len(Warnings) != 0:
        
        for Warning_ in Warnings:
        
            if Warning_ == 'The input contains boundaries below the encastre level.':
                pdf.cell(200, 10, txt = 'The input contains boundaries below the encastre level.', 
                          ln = 6, align = 'L')
                pdf.cell(200, 10, txt = 'Properties below such boundaries are not used in this program version.', 
                          ln = 7, align = 'L')
                pdf.cell(200, 10, txt = 'The effect should be examined by using a boundary immediately above the', 
                          ln = 8, align = 'L')
                pdf.cell(200, 10, txt = 'encastre level with suitably adjusted parameters.', 
                          ln = 9, align = 'L')
            elif Warning_ == 'WinSPOOKS warning: check log file':
                pdf.cell(200, 10, txt = 'WinSPOOKS warning: check log file.', 
                          ln = 6, align = 'L')
                
            elif Warning_ == 'WinSPOOKS error: check log file':
                pdf.cell(200, 10, txt = 'WinSPOOKS error: check log file.', 
                          ln = 6, align = 'L')
    else:
        pdf.cell(200, 10, txt = 'No warnings.', 
                  ln = 6, align = 'L')
    
    pdf.ln(2)
    
    #### GENERAL INFORMATION
    # Header
    pdf.set_font("Courier", size = 15) 
    pdf.cell(200, 10, txt = '1. General information', 
              ln = 10, align = 'L')
    pdf.ln(8)
    pdf.set_font("Courier", size = 12) 
    pdf.cell(200, 10, txt = "1.1 Execution", 
              ln = 11, align = 'L')
    
    # Date
    pdf.cell(col_width1, 2*th, str('Date:'), border=1)
    pdf.cell(col_width2, 2*th, str(Date), border=1)
    pdf.ln(2*th)
    # Project
    pdf.cell(col_width1, 2*th, str('Project:'), border=1)
    pdf.cell(col_width2, 2*th, str(Project), border=1)
    pdf.ln(2*th)
    
    # Initials
    pdf.cell(col_width1, 2*th, str('Initials:'), border=1)
    pdf.cell(col_width2, 2*th, str(Initials), border=1)
    pdf.ln(2*th)
    # Subject
    pdf.cell(col_width1, 2*th, str('Subject:'), border=1)
    pdf.cell(col_width2, 2*th, str(Subject), border=1)
    pdf.ln(2*th)
    # Calc. no.
    pdf.cell(col_width1, 2*th, str('Calculation no.:'), border=1)
    pdf.cell(col_width2, 2*th, str(AnalysisNo), border=1)
    pdf.ln(4*th)
    ## Check
    pdf.cell(200, 10, txt = "1.2 Check", 
              ln = 11, align = 'L')
    pdf.cell(col_width1, 2*th, str('Checker:'), border=1)
    pdf.cell(col_width2, 2*th, str(Checker), border=1)
    pdf.ln(2*th)
    pdf.cell(col_width1, 2*th, str('Date:'), border=1)
    pdf.cell(col_width2, 2*th, str(''), border=1)
    pdf.ln(4*th)
    ## Approval
    pdf.cell(200, 10, txt = "1.2 Approval", 
              ln = 11, align = 'L')
    pdf.cell(col_width1, 2*th, str('Approver:'), border=1)
    pdf.cell(col_width2, 2*th, str(Approver), border=1)
    pdf.ln(2*th)
    # Header
    pdf.ln(2*th)
    pdf.set_font("Courier", size = 15) 
    pdf.cell(200, 10, txt = '2. Input parameters', 
              ln = 14, align = 'L')
    ## Width of columns
    col_width1 = epw/3
    col_width2 = epw/6
    # Top wall
    pdf.set_font("Courier", size = 12)
    pdf.cell(col_width1, 2*th, str('Wall top, zT:'), border=1)
    pdf.cell(col_width2, 2*th, str(zT), border=1)
    pdf.cell(col_width1, 2*th, str('m'), border=1)
    pdf.ln(2*th)

    if AnchorLevel != None:
        # Anchor level
        pdf.cell(col_width1, 2*th, str('Anchor level, zA:'), border=1)
        pdf.cell(col_width2, 2*th, str(AnchorLevel), border=1)
        pdf.cell(col_width1, 2*th, str('m'), border=1)
        pdf.ln(2*th)
        # Anchor inclination
        pdf.cell(col_width1, 2*th, str('Anchor inclination:'), border=1)
        pdf.cell(col_width2, 2*th, str(AnchorInclination), border=1)
        pdf.cell(col_width1, 2*th, str('deg.'), border=1)
        pdf.ln(2*th)
        # Prescribed anchor force
        if PrescrbAnchorForce != 0.00:
            pdf.cell(col_width1, 2*th, str('Prescr. anchor force:'), border=1)
            pdf.cell(col_width2, 2*th, str(PrescrbAnchorForce), border=1)
            pdf.cell(col_width1, 2*th, str('kN/m'), border=1)
            pdf.ln(2*th)
        else:
            pdf.cell(col_width1, 2*th, str('Prescr. anchor force:'), border=1)
            pdf.cell(col_width2, 2*th, str('N/A'), border=1)
            pdf.cell(col_width1, 2*th, str('kN/m'), border=1)
            pdf.ln(2*th)
    # Unit weight of wall
    pdf.cell(col_width1, 2*th, str('Mass of wall:'), border=1)
    pdf.cell(col_width2, 2*th, str(WallMass), border=1)
    pdf.cell(col_width1, 2*th, str('kg/m/m of wall'), border=1)
    pdf.ln(2*th)
    # Density of water
    pdf.cell(col_width1, 2*th, str('Water density, gam_w:'), border=1)
    pdf.cell(col_width2, 2*th, str(WaterDensity), border=1)
    pdf.cell(col_width1, 2*th, str('kN/m3'), border=1)
    pdf.ln(2*th)
    # State
    pdf.cell(col_width1, 2*th, str('State:'), border=1)
    pdf.cell(col_width2, 2*th, str(State), border=1)
    pdf.cell(col_width1, 2*th, str('-'), border=1)
    pdf.ln(2*th)
    # Slope back
    pdf.cell(col_width1, 2*th, str('Slope back:'), border=1)
    pdf.cell(col_width2, 2*th, str(SlopeBack), border=1)
    pdf.cell(col_width1, 2*th, str('deg.'), border=1)
    pdf.ln(2*th)
    # Slope front
    pdf.cell(col_width1, 2*th, str('Slope front:'), border=1)
    pdf.cell(col_width2, 2*th, str(SlopeFront), border=1)
    pdf.cell(col_width1, 2*th, str('deg.'), border=1)
    pdf.ln(2*th)
    # Soil profile
    pdf.cell(col_width1, 2*th, str('Soil profile:'), border=1)
    pdf.cell(col_width2, 2*th, str(SoilProfile), border=1)
    pdf.cell(col_width1, 2*th, str('-'), border=1)
    pdf.ln(2*th)
    pdf.ln(2*th)
    pdf.cell(200, 10, txt = "2.1 Characteristic soil parameters back", 
              ln = 13, align = 'L')
    
    ## Width of columns
    col_width1 = epw/12
    col_width2 = epw/6
    
    
    SoilData = [["z_top", "g_d", "g_m", "cu","c'","phi'","i","r","Description","Keep drained"],["m", "kN/m3", "kN/m3", "kN/m2","kN/m2","deg.","-","-","-","-"]]

    for i in range(0,len(SoilLayersBack)): 
        SoilData.append([format(float(SoilLayersBack[i].get('TopLayer')),'.2f'), 
                         format(float(SoilLayersBack[i].get('Gamma_d')),'.0f'), 
                         format(float(SoilLayersBack[i].get('Gamma_m')),'.0f'), 
                         format(float(SoilLayersBack[i].get('cu')),'.0f'), 
                         format(float(SoilLayersBack[i].get('c')),'.0f'), 
                         format(float(SoilLayersBack[i].get('phi')),'.0f'), 
                         format(float(SoilLayersBack[i].get('i')),'.1f'), 
                         format(float(SoilLayersBack[i].get('r')),'.2f'),
                         SoilLayersBack[i].get('Description'), SoilLayersBack[i].get('KeepDrained')])
                        
    for row in SoilData:
        for cellno, datum in enumerate(row):
            if cellno < 8:
                # Enter data in column
                pdf.cell(col_width1, 2*th, str(datum), border=1)
    
            else:
                pdf.cell(col_width2, 2*th, str(datum), border=1)
     
        pdf.ln(2*th)
    
    pdf.ln(2*th)
    pdf.cell(200, 10, txt = "2.2 Characteristic soil parameters front", 
              ln = 13, align = 'L')
    
    SoilData = [["z_top", "g_d", "g_m", "cu","c'","phi'","i","r","Description","Keep drained"],["m", "kN/m3", "kN/m3", "kN/m2","kN/m2","deg.","-","-","-","-"]]

    for i in range(0,len(SoilLayersFront)): 
        SoilData.append([format(float(SoilLayersFront[i].get('TopLayer')),'.2f'), 
                         format(float(SoilLayersFront[i].get('Gamma_d')),'.0f'), 
                         format(float(SoilLayersFront[i].get('Gamma_m')),'.0f'), 
                         format(float(SoilLayersFront[i].get('cu')),'.0f'), 
                         format(float(SoilLayersFront[i].get('c')),'.0f'), 
                         format(float(SoilLayersFront[i].get('phi')),'.0f'), 
                         format(float(SoilLayersFront[i].get('i')),'.1f'), 
                         format(float(SoilLayersFront[i].get('r')),'.2f'),
                         SoilLayersFront[i].get('Description'), SoilLayersFront[i].get('KeepDrained')])
                        
    
    for row in SoilData:
        for cellno, datum in enumerate(row):
            if cellno < 8:
                # Enter data in colums
                pdf.cell(col_width1, 2*th, str(datum), border=1)
            else:
                pdf.cell(col_width2, 2*th, str(datum), border=1)
     
        pdf.ln(2*th)
    
    ### WATER LEVELS
    col_width = epw/6
    pdf.ln(2*th)
    pdf.cell(200, 10, txt = "2.3 Water levels", 
              ln = 1, align = 'L')
    pdf.cell(col_width, 2*th, str('w_b'), border=1)
    pdf.cell(col_width, 2*th, str('w_f'), border=1)
    pdf.ln(2*th)
    pdf.cell(col_width, 2*th, str('[m]'), border=1)
    pdf.cell(col_width, 2*th, str('[m]'), border=1)
    pdf.ln(2*th)
    pdf.cell(col_width, 2*th, str(WaterLevelBack), border=1)
    pdf.cell(col_width, 2*th, str(WaterLevelFront), border=1)
    pdf.ln(2*th)
    
    #### ADDITIONAL PRESSURE
    pdf.ln(2*th)
    if AddPressureProfile in ['AP1','AP2','AP3','AP4','AP5','AP6','AP7','AP8','AP9','AP10']:
        pdf.cell(200, 10, txt = "2.4 Additional pressure profile:"+' '+AddPressureProfile, 
                  ln = 4, align = 'L')
        
        zAP = ['z [m]']
        for elem in AddPress_z:
            zAP.append(elem)
        eAP = ['ez_k [kN/m2]']
        for elem in AddPress_ez:
            eAP.append(round(elem,1))
            
        eAPd = ['ez_d [kN/m2]']
        for elem in AddPress_ez_Design:
            eAPd.append(round(elem,1))
        
        AP = [zAP,eAP, eAPd]
        
        ## Width of columns
        col_width1 = epw/12
        col_width2 = epw/6
        
            
        for row in AP:
            for cellno, datum in enumerate(row):
                if cellno > 0:
                    # Enter data in colums
                    pdf.cell(col_width1, 2*th, str(datum), border=1)
                else:
                    pdf.cell(col_width2, 2*th, str(datum), border=1)
         
            pdf.ln(2*th)
    else:
        pdf.cell(200, 10, txt = "2.4 Additional pressure profile:"+' '+'no additional pressure', 
                  ln = 4, align = 'L')
    
    #### LOADS
    col_width = epw/6
    pdf.ln(2*th)
    pdf.cell(200, 10, txt = "2.5 Loads", 
              ln = 3, align = 'L')
    pdf.cell(col_width, 2*th, str('zR'), border=1)
    pdf.cell(col_width, 2*th, str('q_bk'), border=1)
    pdf.cell(col_width, 2*th, str('q_fk'), border=1)
    pdf.cell(col_width*2, 2*th, str('Axial wall load (design)'), border=1)
    pdf.ln(2*th)
    pdf.cell(col_width, 2*th, str('m'), border=1)
    pdf.cell(col_width, 2*th, str('kN/m2'), border=1)
    pdf.cell(col_width, 2*th, str('kN/m2'), border=1)
    pdf.cell(col_width*2, 2*th, str('kN/m'), border=1)
    pdf.ln(2*th)
    pdf.cell(col_width, 2*th, str(zR), border=1)
    pdf.cell(col_width, 2*th, str(LoadBack), border=1)
    pdf.cell(col_width, 2*th, str(LoadFront), border=1)
    pdf.cell(col_width*2, 2*th, str(AxialWallLoad), border=1)
    pdf.ln(2*th)
    
    #### SAFETY
    # partial safety factors
    col_width1 = epw*3/12
    col_width2 = epw/12
    pdf.ln(2*th)
    pdf.cell(200, 10, txt = "2.6 Safety", 
              ln = 4, align = 'L')
    pdf.cell(col_width1, 2*th, str('Alpha'), border=1)
    pdf.cell(col_width2, 2*th, str(Alpha), border=1)
    pdf.ln(2*th)
    pdf.cell(col_width1, 2*th, str('Consequence class:'), border=1)
    pdf.cell(col_width2, 2*th, str(ConsequenceClass), border=1)
    pdf.ln(2*th)
    pdf.set_font("Courier", size = 11) 
    psf_header = ['f_gamf','f_qf','f_cf','f_cuf','f_phif','f_wat','f_AP','f_gamb','f_qb','f_cb','f_cub','f_phib']
    for psf in psf_header:
        pdf.cell(col_width2, 2*th, str(psf), border=1)
    pdf.ln(2*th)
    pdf.set_font("Courier", size = 12) 
    psfs = [PartialSafetyFactors.get('f_gamf'),
            PartialSafetyFactors.get('f_qf'),
            PartialSafetyFactors.get('f_cf'),
            PartialSafetyFactors.get('f_cuf'),
            PartialSafetyFactors.get('f_phif'),
            PartialSafetyFactors.get('f_wat'),
            PartialSafetyFactors.get('f_AP'),
            PartialSafetyFactors.get('f_gamb'),
            PartialSafetyFactors.get('f_qb'),
            PartialSafetyFactors.get('f_cb'),
            PartialSafetyFactors.get('f_cub'),
            PartialSafetyFactors.get('f_phib')]
    for psf in psfs:
        pdf.cell(col_width2, 2*th, str(np.round(psf,2)), border=1)
    pdf.ln(2*th)
    
    ### FAILURE MODE
    pdf.ln(2*th)
    pdf.cell(200, 10, txt = "2.7 Failure mode", 
              ln = 5, align = 'L')
    if AnchorLevel != None:
        pdf.cell(col_width2*3, 2*th, 'Anchored wall', border=1)
    else:
        pdf.cell(col_width2*3, 2*th, 'Free wall', border=1)
    pdf.ln(2*th)
    fail_coeff_header = ['iA', 'iB', 'iC']
    fail_coeff = [iA, iB, iC]
    for fail in fail_coeff_header:
        pdf.cell(col_width2, 2*th, str(fail), border=1)
    pdf.ln(2*th)
    
    for fail in fail_coeff:
        pdf.cell(col_width2, 2*th, str(fail), border=1)
    pdf.ln(2*th)
    
    ### KING POST WALL
    pdf.ln(2*th)
    pdf.cell(200, 10, txt = "2.8 King post wall", 
              ln = 6, align = 'L')
    if zB != None and WD != None and CC != None:
        kingpost_header = ['zB', 'WD', 'CC']
        kingpost_unit = ['m','m','m']
        kingpost = [zB, WD, CC]
        for kp in kingpost_header:
            pdf.cell(col_width2, 2*th, str(kp), border=1)
        pdf.ln(2*th)
        
        for kp in kingpost_unit:
            pdf.cell(col_width2, 2*th, str(kp), border=1)
        pdf.ln(2*th)
        
        for kp in kingpost:
            pdf.cell(col_width2, 2*th, str(kp), border=1)
    else:
        pdf.cell(200, 10, txt = "Not king post wall.", 
              ln = 6, align = 'L')
    pdf.ln(2*th)
    
    pdf.add_page('P')
    
    #### Results
    # Header
    pdf.set_font("Courier", size = 15) 
    pdf.cell(200, 10, txt = '3. Results', 
              ln = 6, align = 'L')
    
    pdf.ln(8)
    
    pdf.set_font("Courier", size = 12)
    pdf.cell(200, 10, txt = "3.1 Summary", 
              ln = 6, align = 'L')
    col_width1 = epw*6/12
    col_width2 = epw*2/12
    # Max. moment
    pdf.cell(col_width1, 2*th, str('Max. |moment|:'), border=1)
    pdf.cell(col_width2, 2*th, str(format(abs(MaxMoment),'.1f')), border=1)
    pdf.cell(col_width2, 2*th, str('kNm/m'), border=1)
    pdf.ln(2*th)
    # Max. shear force
    pdf.cell(col_width1, 2*th, str('Max. |shear force|:'), border=1)
    pdf.cell(col_width2, 2*th, str(format(abs(MaxShearForce),'.1f')), border=1)
    pdf.cell(col_width2, 2*th, str('kN/m'), border=1)
    pdf.ln(2*th)
    # Toe level
    pdf.cell(col_width1, 2*th, str('Toe level:'), border=1)
    pdf.cell(col_width2, 2*th, str(format(ToeLevel,'.1f')), border=1)
    pdf.cell(col_width2, 2*th, str('m'), border=1)
    pdf.ln(2*th)

    if AnchorLevel != None:
        # Anchor force
        pdf.cell(col_width1, 2*th, str('Anchor force, Ad:'), border=1)
        pdf.cell(col_width2, 2*th, str(format(AnchorForce,'.1f')), border=1)
        pdf.cell(col_width2, 2*th, str('kN/m'), border=1)
        pdf.ln(2*th)
        # Axial anchor force
        AnchorAxial = AnchorForce/np.cos(np.radians(AnchorInclination))
        pdf.cell(col_width1, 2*th, str('Axial anchor force:'), border=1)
        pdf.cell(col_width2, 2*th, str(format(AnchorAxial,'.1f')), border=1)
        pdf.cell(col_width2, 2*th, str('kN/m'), border=1)
        pdf.ln(2*th)
        # Moment at anchor
        pdf.cell(col_width1, 2*th, str('|Moment| at anchor level:'), border=1)
        pdf.cell(col_width2, 2*th, str(format(abs(float(MomentAtAnchor)),'.1f')), border=1)
        pdf.cell(col_width2, 2*th, str('kNm/m'), border=1)
        pdf.ln(2*th)
    # Tangential earth pressure resultant
    pdf.cell(col_width1, 2*th, str('Sum of tangential earth pressure*:'), border=1)
    pdf.cell(col_width2, 2*th, str(format(SumTanForce,'.1f')), border=1)
    pdf.cell(col_width2, 2*th, str('kN/m'), border=1)
    pdf.ln(2*th)
    # Sum of vertical forces
    pdf.cell(col_width1, 2*th, str('Sum of vertical forces*:'), border=1)
    if AnchorLevel != None:
        pdf.cell(col_width2, 2*th, str(format(SumTanForce-AnchorAxial*np.sin(np.radians(float(AnchorInclination)))-WeightWallTotal,'.1f')), border=1)
    else:
        pdf.cell(col_width2, 2*th, str(format(SumTanForce-WeightWallTotal,'.1f')), border=1)
    pdf.cell(col_width2, 2*th, str('kN/m'), border=1)
    pdf.ln(2*th)
    # Note
    pdf.cell(epw*10/12, 2*th, str('*Tangential pressure and vertical forces are positive upwards.'), border=1)
    pdf.ln(2*th)
    
    ## Pressure and structural forces
    pdf.ln(2*th)
    pdf.cell(200, 10, txt = "3.2 Pressure and structural forces", 
              ln = 6, align = 'L')
    
    col_width = epw/8
    forces_header = ['Level','e1','e2','dw','e-net','Ved','Med','Ju']
    forces_units = ['m','kN/m2','kN/m2','kN/m2','kN/m2','kN/m','kNm/m','-']
    for force in forces_header:
        pdf.cell(col_width, 2*th, str(force), border=1)
    pdf.ln(2*th)
    for force in forces_units:
        pdf.cell(col_width, 2*th, str(force), border=1)
    pdf.ln(2*th)
    for i in range(len(PlotLevels)):
        pdf.cell(col_width, 2*th, str(format(PlotLevels[i],'.1f')), border=1)
        pdf.cell(col_width, 2*th, str(format(e1[i],'.1f')), border=1)
        pdf.cell(col_width, 2*th, str(format(e2[i],'.1f')), border=1)
        pdf.cell(col_width, 2*th, str(format(DW[i],'.1f')), border=1)
        pdf.cell(col_width, 2*th, str(format(ENet[i],'.1f')), border=1)
        pdf.cell(col_width, 2*th, str(format(ShearForce[i],'.1f')), border=1)
        pdf.cell(col_width, 2*th, str(format(Moment[i],'.1f')), border=1)
        pdf.cell(col_width, 2*th, str(format(JU[i],'.1f')), border=1)
        pdf.ln(2*th)
        
    if UseAddOn == 'Yes':
        ## Sheet pile add on
        # Header
        pdf.set_font("Courier", size = 15) 
        pdf.cell(200, 10, txt = '4. Sheet pile add on', 
                  ln = 6, align = 'L')
        
        pdf.ln(2*th)
        pdf.set_font("Courier", size = 12)
        pdf.cell(200, 10, txt = "4.1 Input", 
                  ln = 6, align = 'L')
        
        col_width1 = epw*7/12
        col_width2 = epw*2/12
        # Use add on?
        pdf.cell(col_width1, 2*th, str('Add on active?:'), border=1)
        pdf.cell(col_width2, 2*th, str(UseAddOn), border=1)
        pdf.cell(col_width2, 2*th, str('-'), border=1)
        pdf.ln(2*th)
        # Limit state
        pdf.cell(col_width1, 2*th, str('Limit state:'), border=1)
        pdf.cell(col_width2, 2*th, str(LimitState), border=1)
        pdf.cell(col_width2, 2*th, str('-'), border=1)
        pdf.ln(2*th)
        # Control class
        pdf.cell(col_width1, 2*th, str('Control class:'), border=1)
        pdf.cell(col_width2, 2*th, str(ControlClass), border=1)
        pdf.cell(col_width2, 2*th, str('-'), border=1)
        pdf.ln(2*th)
        # Optimize
        pdf.cell(col_width1, 2*th, str('Optimize:'), border=1)
        pdf.cell(col_width2, 2*th, str(Optimize), border=1)
        pdf.cell(col_width2, 2*th, str('-'), border=1)
        pdf.ln(2*th)
        # Max utilization
        pdf.cell(col_width1, 2*th, str('Max. utilization:'), border=1)
        pdf.cell(col_width2, 2*th, str(MaxUtilization), border=1)
        pdf.cell(col_width2, 2*th, str('-'), border=1)
        pdf.ln(2*th)
        # fyk
        pdf.cell(col_width1, 2*th, str('fyk:'), border=1)
        pdf.cell(col_width2, 2*th, str(fyk), border=1)
        pdf.cell(col_width2, 2*th, str('MPa'), border=1)
        pdf.ln(2*th)
        # Beta B
        pdf.cell(col_width1, 2*th, str('Beta_B:'), border=1)
        pdf.cell(col_width2, 2*th, str(BetaB), border=1)
        pdf.cell(col_width2, 2*th, str('-'), border=1)
        pdf.ln(2*th)
        # Beta D
        pdf.cell(col_width1, 2*th, str('Beta_D:'), border=1)
        pdf.cell(col_width2, 2*th, str(BetaD), border=1)
        pdf.cell(col_width2, 2*th, str('-'), border=1)
        pdf.ln(2*th)
        # Design Life
        pdf.cell(col_width1, 2*th, str('Design life:'), border=1)
        pdf.cell(col_width2, 2*th, str(DesignLife), border=1)
        pdf.cell(col_width2, 2*th, str('Years'), border=1)
        pdf.ln(2*th)
        # Soil compaction
        pdf.cell(col_width1, 2*th, str('Soil compaction:'), border=1)
        pdf.cell(col_width2, 2*th, str(SoilDeposit), border=1)
        pdf.cell(col_width2, 2*th, str('-'), border=1)
        pdf.ln(2*th)
        # corrosion rate at max. M, front
        pdf.cell(epw*11/12, 2*th, str('Corrosion rates (total)'), border=1)
        pdf.ln(2*th)
        pdf.cell(col_width1, 2*th, str('Level (m)'), border=1)
        pdf.cell(col_width2, 2*th, str('Rate'), border=1)
        pdf.cell(col_width2, 2*th, str('Unit'), border=1)
        pdf.ln(2*th)
        for i in range(len(tCor)):
            # total corrosion rates
            pdf.cell(col_width1, 2*th, str(tCorLevel[i]), border=1)
            pdf.cell(col_width2, 2*th, str(tCor[i]), border=1)
            pdf.cell(col_width2, 2*th, str('mm/yr'), border=1)
            pdf.ln(2*th)

        
        
        pdf.ln(2*th)
        pdf.set_font("Courier", size = 12)
        pdf.cell(200, 10, txt = "4.2 Results", 
                  ln = 6, align = 'L')
        col_width1 = epw*6/12
        col_width1a = epw*7/12
        col_width2 = epw*2/12
        col_width2a = epw*5/12
        
        # Sheet pile profile
        pdf.cell(col_width1, 2*th, str('Sheet pile profile:'), border=1)
        pdf.cell(col_width2a, 2*th, str(Sheetpile), border=1)
        pdf.ln(2*th)
        pdf.cell(col_width1a, 2*th, str('Max. relative utilization ratio:'), border=1)
        pdf.cell(col_width2, 2*th, str(u_rel), border=1)
        pdf.cell(col_width2, 2*th, str('-'), border=1)
        pdf.ln(2*th)
        pdf.cell(col_width1a, 2*th, str('Rotational capacity:'), border=1)
        pdf.cell(col_width2*2, 2*th, str(control_Rot), border=1)
        pdf.ln(2*th)
        pdf.cell(col_width1a, 2*th, str('Level (m)'), border=1)
        pdf.cell(col_width2*2, 2*th, str('Rel. utilisation ratio'), border=1)
        pdf.ln(2*th)
        try:
            for index, row in u_rel_lvl.iterrows():
                print(index, row['u_rel'])
                pdf.cell(col_width1a, 2*th, str(index), border=1)
                pdf.cell(col_width2, 2*th, str(round(row['u_rel'],3)), border=1)
                pdf.cell(col_width2, 2*th, str('-'), border=1)
                pdf.ln(2*th)
        except:
            
            pass


      
    # save the pdf with name .pdf

    TemporaryPath = TemporaryWorkingDirectory()
    
    TemporaryPath = os.path.join(TemporaryPath, r'pdfresults.pdf')
    
    ## Save pdf
    pdf.output(TemporaryPath)
    
    
    return TemporaryPath


def ReportGenerator(GetResultsOutput,OutputDirList,Version):
    print('Generating report...')
    
    
    ## Vertical equilibrium
    
    VerticalEquilibriumOutput = VerticalEquilibrium(GetResultsOutput)
    

    if GetResultsOutput['Analysis']['SheetPileAddOnInput']['UseAddOn'] == 'Yes':
       # try:
        SheetPileAddOnResults = steel_sheet_pile_implementer(GetResultsOutput)
        #except Exception as E:
        #    print(f"Error processing sheet pile results: {E}")  # Debugging output
            
        #    SheetPileAddOnResults = {'SheetPileProfile': 'N/A',
        #                             'RUR': 'N/A',
        #                             'RURLevel': 'N/A',
        #                             'RotCap': 'N/A'}
            
    else:
        
        SheetPileAddOnResults = {'SheetPileProfile': 'N/A',
                                 'RUR': 'N/A',
                                 'RURLevel': 'N/A',
                                 'RotCap': 'N/A'}


    
    TempReportFrontPath = ReportFront(VerticalEquilibriumOutput,OutputDirList,Version)
    
    
    TemporaryPathResults = PDFGenerator(VerticalEquilibriumOutput, SheetPileAddOnResults, Version)
    

    
    

    
    ## Temporary file dir
    TemporaryPathReport = TemporaryWorkingDirectory()
    
    ## Temporary report file dir
    TemporaryPathReport = os.path.join(TemporaryPathReport, r'Report.pdf')
                
    
    ### pdfs to merge (existing file, plot, results)               
    pdfsToMerge = [TempReportFrontPath, TemporaryPathResults]
    
    
    ## Merging pdfs
    with contextlib.ExitStack() as stack:
        try:
            pdfMerger = PyPDF2.PdfMerger()
        except AttributeError:
            pdfMerger = PyPDF2.PdfFileMerger()
        files = [stack.enter_context(open(pdf, 'rb')) for pdf in pdfsToMerge]
        for f in files:
            pdfMerger.append(f)
        with open(TemporaryPathReport, 'wb') as f:
            pdfMerger.write(f)
            
        f.close()
    
    ### Deleting input pdfs
    os.remove(TempReportFrontPath)
    os.remove(TemporaryPathResults)

    return TemporaryPathReport
    
    
    
def ReportsMerger(FeederOutput,OutputDirList,Version,stat,tabcalc,pb,calcno):
    
    stat.configure(text = "Generating reports...")
    ## Progress bar maximum
    pb['maximum'] = len(FeederOutput)-1
    pb.update() ## update progress bar
    tabcalc.update_idletasks()
    
    print('Merging report pages...')

    
        
    
    ## Temporary file dir
    TemporaryDir = TemporaryWorkingDirectory()
    
    OutputDirectory = os.path.join(OutputDirList[-1],r'WinSpooksReport.pdf')
    
    ## Inital empty file
    TempFile = os.path.join(TemporaryDir, r'TempFile.pdf')
    TemporaryFile = os.path.join(TemporaryDir, r'WinSpooksReport.pdf')
    
    
    Initial = None
    
    ## Loop through all calculated analyses
    for AnalysisOutput in FeederOutput:
        
        ### PROGRESS BAR AND CALCULATION NO UPDATE
        AnalysisNo = AnalysisOutput.get('ExecuteOutput').get('Analysis').get('AnalysisNo')
        calcno.configure(text = str(AnalysisNo)) ## Calculation number for GUI
        tabcalc.update_idletasks()
        pb['value'] = AnalysisNo
        pb.update() ## update progress bar
        
        
        GetResultsOutput = AnalysisOutput.get('GetResultsOutput')
        
        
        ## Generate report including vertical equilibrium
        
        TemporaryPathReport = ReportGenerator(GetResultsOutput,OutputDirList,Version)
        
        
        if Initial == None:
            
            ## If old report exists -> remove
            try:
                os.remove(TemporaryFile)
                
            except:
                pass
            
            os.rename(TemporaryPathReport,TemporaryFile)
        
        else:
        
            ## Get analysis report (.pdf)
            pdfsToMerge = [TemporaryFile, TemporaryPathReport]
        
            ## Append to initial .pdf file
            ## Merging pdfs
            with contextlib.ExitStack() as stack:
                try:
                    pdfMerger = PyPDF2.PdfMerger()
                except AttributeError:
                    pdfMerger = PyPDF2.PdfFileMerger()
                files = [stack.enter_context(open(pdf, 'rb')) for pdf in pdfsToMerge]
                for f in files:
                    pdfMerger.append(f)
            
                with open(TempFile, 'wb') as f:
                    pdfMerger.write(f)
                        
                    f.close()
               
            try:
                os.remove(TemporaryFile)
            except:
                pass
                
        
            os.rename(TempFile,TemporaryFile)
            
        Initial = 1
            
                
    ## Copy report to user defined destination


    try:
        copyfile(TemporaryFile,OutputDirectory)
    except:
        copyfile(TemporaryPathReport,OutputDirectory)
    
    
    

    
def PlotResults(FeederOutput):
    
    print('Plotting results...')
    
    
    WallResults = []
    

    ## List of parent analysis numbers
    ParentAnalyses = []
    for Analysis in FeederOutput:
        
        GetResultsOutput = Analysis.get('GetResultsOutput')
        ParentAnalyses.append(GetResultsOutput.get('Analysis').get('ParentAnalysis'))
    
    ParentAnalyses = list(np.unique(ParentAnalyses))
    
    for ParentAnalysis in ParentAnalyses:
        
        WallResults.append({'ParentAnalysis': ParentAnalysis,
                            'Subject' : [],
                            'PlotIDs': [],
                            'MaxMoments': [],
                            'MaxShearForces': [],
                            'ToeLevels': [],
                            'AnchorForces': []})
        
        
    for WallResult in WallResults:
        
        ParentAnalysis = WallResult.get('ParentAnalysis')
        
    
        ## ID for plotting
        PlotID = 0
    
        for Analysis in FeederOutput:
            
            GetResultsOutput = Analysis.get('GetResultsOutput')
            
            MaxMomentAnalysis = GetResultsOutput.get('PlotResults').get('MaxMoment')
            
            MaxShearForceAnalysis = GetResultsOutput.get('PlotResults').get('MaxShearForce')
            
            ToeLevel = GetResultsOutput.get('Results').get('ToeLevel')
            
            AnchorForce = GetResultsOutput.get('Results').get('AnchorForce')
            
            ParentAnalysisInput = GetResultsOutput.get('Analysis').get('ParentAnalysis')
            
            Subject = GetResultsOutput.get('Analysis').get('Subject')
            
            if ParentAnalysisInput == ParentAnalysis:
                
                WallResult.get('Subject').append(Subject)
                WallResult.get('PlotIDs').append(PlotID)
                WallResult.get('MaxMoments').append(abs(MaxMomentAnalysis))
                WallResult.get('MaxShearForces').append(abs(MaxShearForceAnalysis))
                WallResult.get('ToeLevels').append(ToeLevel)
                
                if isinstance(AnchorForce,float):
                    
                    WallResult.get('AnchorForces').append(AnchorForce)
                else:
                    WallResult.get('AnchorForces').append(0.00)
            
                ## Updating Plot ID
                PlotID = PlotID + 1
                    


                

    fig = plt.figure()
    #fig.canvas.set_window_title('WinSpooks Plug-in results')
    ### Max. moment
    ax1 = fig.add_subplot(221)
    color = 'black'
    for Results in WallResults:
        plt.plot(Results.get('PlotIDs'),Results.get('MaxMoments'), marker = 'o', alpha = 0.4, label = Results.get('Subject')[0])
    ax1.set_ylabel('Max. |bending moment| [kNm/m]', color = color)
    ax1.tick_params(axis='y', labelcolor = color)
    plt.grid()
    plt.legend()    


    ax2 = fig.add_subplot(222)
    color = 'black'
    for Results in WallResults:
        plt.plot(Results.get('PlotIDs'),Results.get('MaxShearForces'), marker = 'o', alpha = 0.4, label = Results.get('Subject')[0])
    ax2.set_ylabel('Max. |shear force| [kN/m]', color = color)
    ax2.tick_params(axis='y', labelcolor = color)
    plt.grid()
    plt.legend()
    
    ### Anchor force
    ax3 = fig.add_subplot(223)
    color = 'black'
    for Results in WallResults:
        plt.plot(Results.get('PlotIDs'), Results.get('AnchorForces'), marker = 'o', alpha = 0.4, label = Results.get('Subject')[0])
    ax3.set_ylabel('Anchor force [kN/m]', color = color)
    ax3.set_xlabel('Plot ID')
    ax3.tick_params(axis='y', labelcolor = color)
    plt.grid()
    plt.legend()

    ### Toe level
    ax4 = fig.add_subplot(224)
    color = 'black'
    for Results in WallResults:
        plt.plot(Results.get('PlotIDs'), Results.get('ToeLevels'), marker = 'o', alpha = 0.4, label = Results.get('Subject')[0])
    ax4.set_ylabel('Toe level [m]', color = color)
    ax4.set_xlabel('Plot ID')
    ax4.tick_params(axis='y', labelcolor = color)
    plt.grid()
    plt.legend()
    plt.show()

    

def ExportResultsAsTxt(FeederOutput, OutputDirList):
    
    if FeederOutput[0].get('GetResultsOutput')['Analysis']['SheetPileAddOnInput']['UseAddOn'] == "Yes":
        printlist = [['Analysis No.','Subject', 'Soil Profile',
                     'State', 'Load Combination', 'AnchorLevel',
                     'AnchorForce', 'Moment at Anchor Level', 
                     'Level of Max. Moment', 'Max. Moment',
                     'Encastre Level', 'Encastre Moment',
                     'Yield Level', 'Yield Moment', 'Max. Shear Force',
                     'Toe Level', 'Sum of vertical forces','Profile',
                     'Max utilization','Lvl of Max utilization']]
    else:
        printlist = [['Analysis No.','Subject', 'Soil Profile',
                     'State', 'Load Combination', 'AnchorLevel',
                     'AnchorForce', 'Moment at Anchor Level', 
                     'Level of Max. Moment', 'Max. Moment',
                     'Encastre Level', 'Encastre Moment',
                     'Yield Level', 'Yield Moment', 'Max. Shear Force',
                     'Toe Level', 'Sum of vertical forces']]
    
    
    
    for Analysis in FeederOutput:
        GetResultsOutput = Analysis.get('GetResultsOutput')
        
        print(GetResultsOutput)
        
        if GetResultsOutput['Analysis']['SheetPileAddOnInput']['UseAddOn'] == 'Yes':
            
            try:
                SheetPileAddOnResults = steel_sheet_pile_implementer(GetResultsOutput)
            except Exception as e:
                print(f"Error processing sheet pile results: {e}")
                SheetPileAddOnResults = {'SheetPileProfile': 'N/A',
                                         'RUR': 'N/A',
                                         'RURLevel': 'N/A',
                                         'RURLevel_max': 'N/A',
                                         'RotCap': 'N/A'}                
        else:
            
            SheetPileAddOnResults = {'SheetPileProfile': 'N/A',
                                     'RUR': 'N/A',
                                     'RURLevel': 'N/A',
                                     'RURLevel_max': 'N/A',
                                     'RotCap': 'N/A'}
        
        ## Vertical equilibrium
        VerticalEquilibriumOutput = VerticalEquilibrium(GetResultsOutput)
        
        AnalysisNo = GetResultsOutput.get('Analysis').get('AnalysisNo')
        zT = GetResultsOutput.get('Analysis').get('zT')
        Subject = GetResultsOutput.get('Analysis').get('Subject')
        SoilProfile = GetResultsOutput.get('Analysis').get('SoilProfile')
        LoadCombination = GetResultsOutput.get('Analysis').get('LoadCombination')
        State = GetResultsOutput.get('Analysis').get('State')
        AxialWallLoad = GetResultsOutput.get('Analysis').get('AxialWallLoad')
        try: 
            AnchorLevel = GetResultsOutput.get('Analysis').get('AnchorLevel')
            AnchorInclination = GetResultsOutput.get('Analysis').get('AnchorInclination')
            AnchorForce = GetResultsOutput.get('Results').get('AnchorForce')
            AnchorAxial = AnchorForce/np.cos(np.radians(AnchorInclination))
            MomentAtAnchor = GetResultsOutput.get('Results').get('MomentAtAnchor')
        except:
            AnchorLevel = 'N/A'
            AnchorInclination = 'N/A'
            AnchorForce = 'N/A'
            AnchorAxial = 'N/A'
            MomentAtAnchor = 'N/A'
        LevelMaxMoment = GetResultsOutput.get('PlotResults').get('LevelMaxMoment')
        MaxMoment = GetResultsOutput.get('PlotResults').get('MaxMoment')
        EncastreLevel = GetResultsOutput.get('Results').get('EncastreLevel')
        EncastreMoment = GetResultsOutput.get('Results').get('EncastreMoment')
        YieldHingeLevel = GetResultsOutput.get('Results').get('YieldHingeLevel')
        YieldMoment = GetResultsOutput.get('Results').get('YieldHingeMoment')
        MaxShearForce = GetResultsOutput.get('PlotResults').get('MaxShearForce')
        ToeLevel = GetResultsOutput.get('Results').get('ToeLevel')
        if ToeLevel == "N/A":
            DesignSoilLayersFront = GetResultsOutput.get('Analysis').get('SoilLayersFront')
            SoilLevelsFront = []
            for SoilLayer in DesignSoilLayersFront:
                SoilLevelsFront.append(SoilLayer.get('TopLayer'))
            GroundLevelFront = max(SoilLevelsFront)
            ToeLevel = GroundLevelFront - 0.1
        SumTanForce = VerticalEquilibriumOutput.get('SumTanForce')
        WallMass = GetResultsOutput.get('Analysis').get('WallMass')
        WallSpecificWeight = WallMass * 9.82 / 1000
        WeightWallTotal = WallSpecificWeight * (zT - ToeLevel)
        
        sel_profile = SheetPileAddOnResults.get('SheetPileProfile')
        max_util = SheetPileAddOnResults.get('RUR')
        max_util_lvl = SheetPileAddOnResults.get('RURLevel_max')
        
        
        if AnchorLevel != 'N/A':
            SumVertForces = SumTanForce-AnchorAxial*np.sin(np.radians(float(AnchorInclination)))-WeightWallTotal-AxialWallLoad
        else:
            SumVertForces =  SumTanForce-WeightWallTotal-AxialWallLoad
        
    
    #### Generating analyses results list
    
        if GetResultsOutput['Analysis']['SheetPileAddOnInput']['UseAddOn'] == 'Yes':
            printlist_temp = [AnalysisNo, Subject, SoilProfile, 
                              State, LoadCombination, AnchorLevel, 
                              AnchorForce, MomentAtAnchor, LevelMaxMoment, 
                              MaxMoment, EncastreLevel, EncastreMoment, 
                              YieldHingeLevel, YieldMoment, MaxShearForce, 
                              ToeLevel, SumVertForces, sel_profile, max_util,
                              max_util_lvl]
        else:
            printlist_temp = [AnalysisNo, Subject, SoilProfile, 
                              State, LoadCombination, AnchorLevel, 
                              AnchorForce, MomentAtAnchor, LevelMaxMoment, 
                              MaxMoment, EncastreLevel, EncastreMoment, 
                              YieldHingeLevel, YieldMoment, MaxShearForce, 
                              ToeLevel, SumVertForces]

        printlist.append(printlist_temp)
    
    results_path = OutputDirList[-1]
    results_path = os.path.join(results_path, r'Results.txt')
    np.savetxt(results_path, printlist, delimiter=';', fmt='%s')
        
        ###############################
    


def ExportEarthPressureResultsAsTxt(FeederOutput, OutputDirList):
    
    ## Temporary file dir
    #TemporaryFile = TemporaryWorkingDirectory()
    ## Initial empty array
    EarthPressureResults = np.empty((0,8), int) # Empty array
    
    for Analysis in FeederOutput:

        #GetResultsOutput = Analysis.get('GetResultsOutput')
        
        EarthPressureResultsOutput = Analysis.get('GetResultsOutput').get('EarthPressureResults')
        
        try:
            AnalysisNo = 'Analysis no. '+ Analysis.get('AnalysisNo')
        except:
            AnalysisNo = 'N/A'
        AnalysisNoArray = np.empty((0,8), int) ## empty array
        np.append(AnalysisNoArray, ['AnalysisNo', AnalysisNo,'','','','','',''])
        
    
        EarthPressureResults = np.concatenate((EarthPressureResults, AnalysisNoArray, EarthPressureResultsOutput), axis=0)
        
    
    
    # Exporting earth pressure results as text file
    results_path = OutputDirList[-1] ## results path
    EP_results = os.path.join(results_path, r'EarthPressureResults.txt')
    np.savetxt(EP_results, EarthPressureResults, delimiter='\t', fmt='%s')
    
    
    return EarthPressureResults

              





############### GUI Functions

def OpenSpooks():
    
    ################### OPEN WINSPOOKS ###################################
    ######### Check in WinSpooks is running - if not -> Run
    ######### (necessary for license check)

    res = subprocess.check_output(['tasklist']).splitlines()
    winspooks = []
    for i in range(0,len(res)):
        p = str(res[i])
        if 'WinSpooks.exe' in p:
            winspooks.append(i) 
    
    if winspooks == []:

        subprocess.Popen(['C:\Program Files (x86)\WinSpooks\WinSpooks.exe'])
        time.sleep(0.5)
    
    res = None  ## Deleting list
    
    


################### 1. OPEN WINSPOOKS ###################################
######### Check in WinSpooks is running - if not -> Run
######### (necessary for license check)


OpenSpooks()
    

################# 2. STARTING UP GUI INTERFACE #######################




window = tk.Tk()
window.minsize(1300,600) ## Initial GUI size
window.title("WinSPOOKS Plug-in")

tab_parent = ttk.Notebook(window)  ## parent tab
tabcalc = ttk.Frame(tab_parent)    ## Calculation tab
tablog = ttk.Frame(tab_parent)     ## Log tab
tabvers = ttk.Frame(tab_parent)    ## Version tab
tab_parent.add(tabcalc, text="Calculations")
tab_parent.add(tablog, text = "Calculation log")
tab_parent.add(tabvers, text="Version")



######## Output file path
outputdir = None
OutputDirList = []

####### Excel input file path
filename = None
FilenameList = []


def FileDialog():
    
    filename = filedialog.askopenfilename(initialdir =  "/", title = "Select A File", filetype =
    (("excel","*.xlsx"),("all files","*.*")))
    FilenameList.append(filename)
    label_file.configure(text = filename)
    label_file.configure(foreground = 'green')
    stat.configure(text = "Input file loaded")
    button_calc.configure(stat = 'normal')
        
    if checkexp.get() == 1 and OutputDirList != [] and FilenameList != []:
        stat.configure(text = "Ready")
    if checkexp.get() == 0 and FilenameList != []:
        stat.configure(text = "Ready")


def OutputDialog():
    
    
    outputdir = filedialog.askdirectory()
    OutputDirList.append(os.path.abspath(outputdir))
    label_dir.configure(text = outputdir)
    label_dir.configure(foreground = 'green')
    
    if checkexp.get() == 1 and OutputDirList != [] and FilenameList != []:
        stat.configure(text = "Ready")

###### Configuration of GUI output browse button    
def Disable():
    if checkexp.get() == 1 or checkexp_ep.get() == 1 or checkreport.get() == 1:
        button_output.configure(stat = 'normal')
        label_dir.configure(foreground = 'black')
        label_outputdir.configure(foreground = 'black')
    else:
        button_output.configure(stat = 'disable')
        label_dir.configure(foreground = 'grey')
        label_outputdir.configure(foreground = 'grey')


## Calculate
def Calculate():
    
    log_usage()


    if checkexp.get() == 1 and OutputDirList == []:
        stat.configure(text = "Choose export directory")
        
        
    elif FilenameList != []:
        
        InputFile = FilenameList[-1]
                
        ## Update GUI
        stat.configure(text = "Calculating...")
        button_calc.configure(stat = 'disable')
        tabcalc.update_idletasks()
        
        ## Run SpooksFeeder
        print("Feeder")
        FeederOutput = SPOOKSFeeder(InputFile,calcno,pb,tabcalc,logtxt,tk)
        
        stat.configure(text = "Calculation completed")
        tabcalc.update_idletasks()
        
        
        ##### Generating reports
        if checkreport.get() == 1:
            
            
            ReportsMerger(FeederOutput,OutputDirList,Version,stat,tabcalc,pb,calcno)
            
            stat.configure(text = "Reports generated")
            tabcalc.update_idletasks()
            
        
        ## Plot results
        if checkplot.get() == 1:
            
            stat.configure(text = "Plotting results...")
            tabcalc.update_idletasks()
            
            PlotResults(FeederOutput)
            
        ## Export results
        if checkexp.get() == 1:
            
            stat.configure(text = "Exporting results...")
            tabcalc.update_idletasks()
            
            ExportResultsAsTxt(FeederOutput,OutputDirList)
            
        
        ## Export earth pressure results
        if checkexp_ep.get() == 1:
            
            stat.configure(text = "Exporting results...")
            tabcalc.update_idletasks()
            
            ExportEarthPressureResultsAsTxt(FeederOutput, OutputDirList)
            
        ## Enable calculation button at end of tasks
        stat.configure(text = "Tasks completed...")
        button_calc.configure(stat = 'normal')
        tabcalc.update_idletasks()

           

####################### GUI ELEMENTS ###############################
####################### Calculation tab #############################
        
### Logo and header
label_logo = ttk.Label(tabcalc, text = " COWI", font=("Century Schoolbook", 33), foreground = 'orange red')
label_logo.grid(column = 0, row = 0, sticky = 'SE')
label_header = ttk.Label(tabcalc, text = "WinSPOOKS Plug-in", font=("Calibri", 16))
label_header.grid(column = 2, row = 0, sticky = 'SW')
## empty row
# enter0 = ttk.Label(tabcalc, text = "",)
# enter0.grid(column = 0, row = 1)
label_bestpractice = ttk.Label(tabcalc, text = " Best practice ", font=("Calibri", 10, 'italic'))
label_bestpractice.grid(column = 0, row = 1, sticky = 'NE')
header_input = ttk.Label(tabcalc, text = "Input", font = ("Calibri",14))
header_input.grid(column = 0, row = 2, sticky = 'E')
button_browse = ttk.Button(tabcalc, text = "Browse", command = FileDialog)
button_browse.grid(column =0, row = 3, sticky = 'E')

label_input = ttk.Label(tabcalc, text = "Input file:  ",font = ("Calibri",11))
label_input.grid(column = 1, row = 3, sticky = 'E')
label_file = ttk.Label(tabcalc, text = "No file",font = ("Calibri",11))
label_file.grid(column = 2, row = 3, sticky = 'W')
enter1 = ttk.Label(tabcalc, text = "_______________")
enter1.grid(column = 0, row = 4, sticky = 'E')

#### empty cell
enter2 = ttk.Label(tabcalc, text = "")
enter2.grid(column = 0, row = 7, sticky = 'E')
#### header output
header_output = ttk.Label(tabcalc, text = "Output", font = ("Calibri",14))
header_output.grid(column = 0, row = 8, sticky = 'E')
## Checkbox plot
checkplot = tk.IntVar()
checkplot.set(0)
checkBox_plot = ttk.Checkbutton(tabcalc, variable=checkplot, onvalue=1, offvalue=0, text="Plot results")
checkBox_plot.grid(column = 1, row = 9, sticky = 'W')
## Checkbox export earth pressure results
checkexp_ep = tk.IntVar()
checkexp_ep.set(0)
checkBox_exportep = ttk.Checkbutton(tabcalc, variable=checkexp_ep, onvalue=1, offvalue=0, text="Export earth pressure results as .txt",command = Disable)
checkBox_exportep.grid(column = 1, row = 10, sticky = 'W')
## Checkbox export structure results
checkexp = tk.IntVar()
checkexp.set(0)
checkBox_export = ttk.Checkbutton(tabcalc, variable=checkexp, onvalue=1, offvalue=0, text="Export results as .txt",command = Disable)
checkBox_export.grid(column = 1, row = 11, sticky = 'W')
## Checkbox generate reports
checkreport = tk.IntVar()
checkreport.set(0)
checkBox_report = ttk.Checkbutton(tabcalc, variable=checkreport, onvalue=1, offvalue=0, text="Generate reports",command = Disable)
checkBox_report.grid(column = 1, row = 12, sticky = 'W')
## text export results
label_outputdir = ttk.Label(tabcalc, text ="Export results to:  ",font = ("Calibri",11), foreground = 'grey')
label_outputdir.grid(column = 1, row = 13, sticky = 'E')
## directory path
label_dir = ttk.Label(tabcalc, text = "No directory",font = ("Calibri",11), foreground = 'grey')
label_dir.grid(column = 2, row = 13, sticky = 'W')

## Browse button
button_output = ttk.Button(tabcalc, text = "Browse", command = OutputDialog, state='disabled')
button_output.grid(column =0, row = 13, sticky = 'E')
#### empty cell
enter3 = ttk.Label(tabcalc, text = "_______________")
enter3.grid(column = 0, row = 14, sticky = 'E')
#### empty cell
enter4 = ttk.Label(tabcalc, text = "________________________________________")
enter4.grid(column = 1, row = 14, sticky = 'W')
#### empty cell
enter5 = ttk.Label(tabcalc, text = "")
enter5.grid(column = 0, row = 15, sticky = 'W')
#### header calculations
header_calc = ttk.Label(tabcalc, text = "Calculations", font = ("Calibri",14))
header_calc.grid(column = 0, row = 16, sticky = 'E')
## status label
label_stat = ttk.Label(tabcalc, text = "Status:",font = ("Calibri",11))
label_stat.grid(column = 1, row = 17, sticky = 'E')
## status 
stat = ttk.Label(tabcalc, text = " No input file",font = ("Calibri",11))
stat.grid(column = 2, row = 17, sticky = 'W')
## Calculation number label
label_calcno = ttk.Label(tabcalc, text = "Calc. no.:" ,font = ("Calibri",11))
label_calcno.grid(column = 1, row = 18, sticky = 'E')
## Calculation number 
calcno = ttk.Label(tabcalc, text = "" ,font = ("Calibri",11))
calcno.grid(column = 2, row = 18, sticky = 'W')
#### empty cell
enter6 = ttk.Label(tabcalc, text = "")
enter6.grid(column = 1, row = 19)
## Calculation button
button_calc = ttk.Button(tabcalc, text="Run calculations", command=Calculate)
button_calc.grid(column = 1, row = 20, sticky = 'E')
## Warning label
warning = ttk.Label(tabcalc, text = "", font = ("Calibri",11),foreground = 'red')
warning.grid(column = 2, row = 20, sticky = 'W')
## Progressbar
pb = ttk.Progressbar(tabcalc, length=300, mode='determinate')
pb.grid(column = 2, row = 21, sticky = 'W')

#### empty cell
enter7 = ttk.Label(tabcalc, text = "")
enter7.grid(column = 0, row = 21)
#### empty cell
enter8 = ttk.Label(tabcalc, text = "")
enter8.grid(column = 0, row = 22)
#### empty cell
enter9 = ttk.Label(tabcalc, text = "")
enter9.grid(column = 0, row = 23)
#### empty cell
enter10 = ttk.Label(tabcalc, text = "")
enter10.grid(column = 0, row = 24)
#### empty cell
enter11 = ttk.Label(tabcalc, text = "")
enter11.grid(column = 0, row = 25)
## COWI A/S
label_cowi = ttk.Label(tabcalc, text = "COWI A/S",font = ("Calibri",9))
label_cowi.grid(column = 0, row = 26, sticky = 'E')
## Log
text = 'Current version: '+Version
label_vers = ttk.Label(tabcalc, text = text,font = ("Calibri",9))
label_vers.grid(column = 2, row = 26, sticky = 'W')

####################### End of Calculation tab #############################

####################### Log tab #########################################
### Logo and header
label_logo = ttk.Label(tablog, text = " COWI", font=("Century Schoolbook", 33), foreground = 'orange red')
label_logo.grid(column = 0, row = 0, sticky = 'SE')
label_header = ttk.Label(tablog, text = "Calculation log", font=("Calibri", 16))
label_header.grid(column = 1, row = 0, sticky = 'S')

logtxt = tkst.ScrolledText(tablog, width=70, height=25)
logtxt.insert(tk.END,'Load file - start calculating.')
logtxt.grid(column=1, row=1)
####################### End of log tab #############################

####################### Version tab #########################################
### Logo and header
label_logo = ttk.Label(tabvers, text = " COWI", font=("Century Schoolbook", 33), foreground = 'orange red')
label_logo.grid(column = 0, row = 0, sticky = 'SE')
label_header = ttk.Label(tabvers, text = "Version", font=("Calibri", 16))
label_header.grid(column = 4, row = 0, sticky = 'SW')
#### empty cell
enter11 = ttk.Label(tabvers, text = "")
enter11.grid(column = 1, row = 1)
#### empty cell
enter12 = ttk.Label(tabvers, text = "")
enter12.grid(column = 1, row = 2)
label_vershead_vers = ttk.Label(tabvers, text = "Version",font = ("Calibri",12))
label_vershead_vers.grid(column = 1, row = 3, sticky = 'W')
label_vers_vers2 = ttk.Label(tabvers, text = "2.0",font = ("Calibri",12))
label_vers_vers2.grid(column = 1, row = 4, sticky = 'W')
label_vers_vers1 = ttk.Label(tabvers, text = "1.0",font = ("Calibri",12))
label_vers_vers1.grid(column = 1, row = 5, sticky = 'W')
label_vers_vers2 = ttk.Label(tabvers, text = "0.1",font = ("Calibri",12))
label_vers_vers2.grid(column = 1, row = 6, sticky = 'W')
label_vershead_date = ttk.Label(tabvers, text = "Date",font = ("Calibri",12))
label_vershead_date.grid(column = 2, row = 3, sticky = 'W')
label_vers_date1 = ttk.Label(tabvers, text = "22.03.2024",font = ("Calibri",12))
label_vers_date1.grid(column = 2, row = 4, sticky = 'W')
label_vers_date1 = ttk.Label(tabvers, text = "12.05.2020",font = ("Calibri",12))
label_vers_date1.grid(column = 2, row = 5, sticky = 'W')
label_vers_date2 = ttk.Label(tabvers, text = "10.01.2020",font = ("Calibri",12))
label_vers_date2.grid(column = 2, row = 6, sticky = 'W')
label_vershead_prep = ttk.Label(tabvers, text = "Prepared",font = ("Calibri",12))
label_vershead_prep.grid(column = 3, row = 3, sticky = 'W')
label_vers_prep1 = ttk.Label(tabvers, text = "EMBT",font = ("Calibri",12))
label_vers_prep1.grid(column = 3, row = 4, sticky = 'W')
label_vers_prep1 = ttk.Label(tabvers, text = "EMBT",font = ("Calibri",12))
label_vers_prep1.grid(column = 3, row = 5, sticky = 'W')
label_vers_prep2 = ttk.Label(tabvers, text = "EMBT",font = ("Calibri",12))
label_vers_prep2.grid(column = 3, row = 6, sticky = 'W')
label_vershead_check = ttk.Label(tabvers, text = "Checked",font = ("Calibri",12))
label_vershead_check.grid(column = 4, row = 3, sticky = 'W')
label_vers_check1 = ttk.Label(tabvers, text = "MLHU",font = ("Calibri",12))
label_vers_check1.grid(column = 4, row = 4, sticky = 'W')
label_vers_check1 = ttk.Label(tabvers, text = "EMSS",font = ("Calibri",12))
label_vers_check1.grid(column = 4, row = 5, sticky = 'W')
label_vers_check2 = ttk.Label(tabvers, text = "EMSS",font = ("Calibri",12))
label_vers_check2.grid(column = 4, row = 6, sticky = 'W')
label_vershead_desc = ttk.Label(tabvers, text = "Description",font = ("Calibri",12))
label_vershead_desc.grid(column = 5, row = 3, sticky = 'W')
label_vers_desc1 = ttk.Label(tabvers, text = "Added sheet pile wall - Add on",font = ("Calibri",12))
label_vers_desc1.grid(column = 5, row = 4, sticky = 'W')
label_vers_desc1 = ttk.Label(tabvers, text = "Report generation and vertical equilibrium added",font = ("Calibri",12))
label_vers_desc1.grid(column = 5, row = 5, sticky = 'W')
label_vers_desc2 = ttk.Label(tabvers, text = "Test version",font = ("Calibri",12))
label_vers_desc2.grid(column = 5, row = 6, sticky = 'W')
####################### End of log tab #############################

tab_parent.pack(expand=1, fill='both')

window.mainloop()